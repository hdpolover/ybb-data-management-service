#!/usr/bin/env python3
"""
Test script to verify export metrics are working correctly
"""
import sys
import os
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from services.ybb_export_service import YBBExportService
import json
import time

def test_export_metrics():
    """Test export timing and file size metrics"""
    
    print("🔧 Testing Export Metrics")
    print("=" * 40)
    
    # Create test data of different sizes
    test_cases = [
        {
            "name": "Small Export (5 records)",
            "data": [
                {
                    'id': i,
                    'full_name': f'Test User {i}',
                    'email': f'test{i}@example.com',
                    'status': 'Approved',
                    'registration_date': '2025-07-27',
                    'country': 'Japan',
                    'institution': 'Test University',
                    'phone_number': f'+81-90-1234-{i:04d}',
                    'category': 'Student'
                } for i in range(1, 6)
            ]
        },
        {
            "name": "Medium Export (50 records)",
            "data": [
                {
                    'id': i,
                    'full_name': f'Participant {i:03d}',
                    'email': f'participant{i}@university.edu',
                    'status': 'Approved' if i % 2 == 0 else 'Pending',
                    'registration_date': '2025-07-27',
                    'country': 'Japan' if i % 3 == 0 else 'Korea',
                    'institution': f'University {i % 10}',
                    'phone_number': f'+{81 if i % 2 == 0 else 82}-{i:02d}-{i*100:04d}',
                    'category': 'Student' if i % 3 == 0 else 'Professional',
                    'notes': f'Additional notes for participant {i} with some longer text content'
                } for i in range(1, 51)
            ]
        }
    ]
    
    export_service = YBBExportService()
    
    for test_case in test_cases:
        print(f"\n📊 {test_case['name']}")
        print("-" * 30)
        
        export_request = {
            'export_type': 'participants',
            'template': 'standard',
            'data': test_case['data'],
            'format': 'excel'
        }
        
        try:
            # Measure total time including API overhead
            api_start = time.time()
            
            result = export_service.create_export(export_request)
            
            api_time = round((time.time() - api_start) * 1000, 2)
            
            if result['status'] == 'success':
                export_id = result['data']['export_id']
                
                # Print export metrics
                print(f"✅ Export Created: {export_id[:8]}...")
                print(f"📈 Records: {result['data']['record_count']}")
                print(f"📁 File Size: {result['data']['file_size']:,} bytes ({result['data'].get('file_size_mb', 0)} MB)")
                print(f"⏱️  Processing Time: {result['metadata']['processing_time_ms']} ms ({result['metadata']['processing_time_seconds']} sec)")
                print(f"🚀 Records/Second: {result['metadata']['records_per_second']}")
                print(f"⏰ Total API Time: {api_time} ms")
                
                # Test status endpoint
                status = export_service.get_export_status(export_id)
                print(f"📋 Status Check:")
                print(f"   File Size: {status.get('file_size_bytes', 'N/A'):,} bytes")
                print(f"   Processing: {status.get('processing_time_ms', 'N/A')} ms")
                print(f"   Rate: {status.get('records_per_second', 'N/A')} rec/sec")
                
                # Test file download
                file_content, filename = export_service.download_export(export_id, 'single')
                if file_content:
                    actual_size = len(file_content)
                    print(f"📥 Download:")
                    print(f"   Filename: {filename}")
                    print(f"   Actual Size: {actual_size:,} bytes")
                    print(f"   Size Match: {'✅' if actual_size == result['data']['file_size'] else '❌'}")
                    
                    # Validate file
                    from openpyxl import load_workbook
                    from io import BytesIO
                    
                    wb = load_workbook(BytesIO(file_content))
                    ws = wb.active
                    print(f"📋 File Validation:")
                    print(f"   Rows: {ws.max_row}")
                    print(f"   Expected Rows: {len(test_case['data']) + 1}")  # +1 for header
                    print(f"   Validation: {'✅' if ws.max_row == len(test_case['data']) + 1 else '❌'}")
                    wb.close()
                
            else:
                print(f"❌ Export Failed: {result.get('message')}")
                
        except Exception as e:
            print(f"❌ Test Failed: {str(e)}")
    
    print(f"\n🔄 Testing Large Export (Chunked)")
    print("-" * 30)
    
    # Test chunked export with larger dataset
    large_data = [
        {
            'id': i,
            'full_name': f'Large Dataset User {i:05d}',
            'email': f'user{i}@largetest.com',
            'status': 'Approved',
            'registration_date': '2025-07-27',
            'country': f'Country {i % 20}',
            'institution': f'Institution {i % 50}',
            'phone_number': f'+{i % 100}-{i:05d}',
            'category': 'Bulk Import',
            'notes': f'Large dataset entry {i} with extended information for testing chunked exports'
        } for i in range(1, 1001)  # 1000 records
    ]
    
    large_export_request = {
        'export_type': 'participants',
        'template': 'standard',
        'data': large_data,
        'format': 'excel'
    }
    
    try:
        api_start = time.time()
        large_result = export_service.create_export(large_export_request)
        api_time = round((time.time() - api_start) * 1000, 2)
        
        if large_result['status'] == 'success':
            print(f"✅ Large Export Strategy: {large_result.get('export_strategy', 'single_file')}")
            
            if large_result.get('export_strategy') == 'multi_file':
                # Multi-file metrics
                archive_info = large_result['data']['archive']
                metadata = large_result['metadata']
                
                print(f"📈 Records: {large_result['data']['total_records']:,}")
                print(f"📁 Files: {large_result['data']['total_files']}")
                print(f"🗜️  ZIP Size: {archive_info['zip_file_size']:,} bytes ({archive_info['zip_file_size_mb']} MB)")
                print(f"📊 Uncompressed: {archive_info['total_uncompressed_size']:,} bytes ({archive_info['total_uncompressed_size_mb']} MB)")
                print(f"📉 Compression: {archive_info['compression_ratio_percent']}%")
                print(f"⏱️  Processing: {metadata['processing_time_ms']} ms ({metadata['processing_time_seconds']} sec)")
                print(f"🚀 Records/Sec: {metadata['records_per_second']}")
                print(f"📁 Files/Sec: {metadata['files_per_second']}")
                print(f"💾 Space Saved: {metadata['space_saved_mb']} MB")
            else:
                # Single file metrics
                print(f"📈 Records: {large_result['data']['record_count']:,}")
                print(f"📁 File Size: {large_result['data']['file_size']:,} bytes ({large_result['data']['file_size_mb']} MB)")
                print(f"⏱️  Processing: {large_result['metadata']['processing_time_ms']} ms")
                print(f"🚀 Records/Sec: {large_result['metadata']['records_per_second']}")
            
            print(f"⏰ Total API Time: {api_time} ms")
        else:
            print(f"❌ Large Export Failed: {large_result.get('message')}")
    
    except Exception as e:
        print(f"❌ Large Export Test Failed: {str(e)}")
    
    print(f"\n✅ Metrics Testing Complete")

if __name__ == "__main__":
    test_export_metrics()
