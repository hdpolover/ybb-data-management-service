"""
Excel export utilities for YBB Data Management API
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from io import BytesIO
import logging
import re
import unicodedata

logger = logging.getLogger(__name__)

class ExcelExporter:
    """Handle Excel export operations with advanced formatting and data sanitization"""
    
    @staticmethod
    def sanitize_cell_value(value):
        """
        Sanitize cell values for Excel compatibility
        
        Args:
            value: Raw cell value
            
        Returns:
            Sanitized value safe for Excel
        """
        if value is None:
            return ""
        
        # Convert to string
        str_value = str(value)
        
        # Handle empty strings
        if not str_value.strip():
            return ""
        
        # Handle Excel formula injection (security)
        if str_value.strip().startswith(('=', '+', '-', '@')):
            str_value = "'" + str_value  # Prefix with apostrophe to treat as text
        
        # Remove or replace problematic characters for Excel
        # Excel doesn't support characters below ASCII 32 except tab (9), newline (10), carriage return (13)
        cleaned_value = ""
        for char in str_value:
            char_code = ord(char)
            if char_code < 32:
                if char_code in [9, 10, 13]:  # Tab, newline, carriage return
                    cleaned_value += char
                else:
                    # Replace with space for other control characters
                    cleaned_value += " "
            elif char_code == 127:  # DEL character
                cleaned_value += " "
            elif char_code > 1114111:  # Beyond valid Unicode range
                cleaned_value += " "
            else:
                cleaned_value += char
        
        # Normalize Unicode characters for better compatibility
        try:
            cleaned_value = unicodedata.normalize('NFKC', cleaned_value)
        except Exception:
            # If normalization fails, try to encode/decode to remove problematic characters
            try:
                cleaned_value = cleaned_value.encode('utf-8', errors='ignore').decode('utf-8')
            except Exception:
                # Last resort: keep only ASCII printable characters
                cleaned_value = ''.join(char for char in str_value if 32 <= ord(char) <= 126)
        
        # Remove excessive whitespace but preserve single spaces and line breaks
        cleaned_value = re.sub(r'[ \t]+', ' ', cleaned_value)  # Multiple spaces/tabs to single space
        cleaned_value = re.sub(r'\n+', '\n', cleaned_value)    # Multiple newlines to single
        cleaned_value = cleaned_value.strip()
        
        # Excel cell limit (32,767 characters)
        if len(cleaned_value) > 32767:
            cleaned_value = cleaned_value[:32764] + "..."
        
        # Final check: ensure it's not empty after cleaning
        if not cleaned_value:
            return ""
        
        return cleaned_value
    
    @staticmethod
    def sanitize_dataframe(df):
        """
        Sanitize entire DataFrame for Excel compatibility
        
        Args:
            df: Pandas DataFrame
            
        Returns:
            Sanitized DataFrame
        """
        df_copy = df.copy()
        
        # Sanitize column names
        sanitized_columns = []
        for col in df_copy.columns:
            sanitized_col = ExcelExporter.sanitize_cell_value(col)
            # Ensure column name is not empty
            if not sanitized_col:
                sanitized_col = "Column"
            sanitized_columns.append(sanitized_col)
        
        df_copy.columns = sanitized_columns
        
        # Sanitize all cell values
        for col in df_copy.columns:
            df_copy[col] = df_copy[col].apply(ExcelExporter.sanitize_cell_value)
        
        return df_copy
    
    @staticmethod
    def create_excel_file(data, filename=None, sheet_name="Data", format_options=None):
        """
        Create Excel file from data with formatting options and data sanitization
        
        Args:
            data: List of dictionaries or pandas DataFrame
            filename: Output filename
            sheet_name: Excel sheet name
            format_options: Dictionary with formatting options
        
        Returns:
            BytesIO object containing Excel file
        """
        try:
            # Convert to DataFrame if needed
            if isinstance(data, list):
                df = pd.DataFrame(data)
            else:
                df = data.copy()
            
            if df.empty:
                raise ValueError("No data to export")
            
            # Sanitize the DataFrame
            df = ExcelExporter.sanitize_dataframe(df)
            
            # Sanitize sheet name
            sheet_name = ExcelExporter.sanitize_cell_value(sheet_name)
            if not sheet_name:
                sheet_name = "Data"
            
            # Excel sheet name limitations
            sheet_name = re.sub(r'[\\/*\[\]:?]', '_', sheet_name)  # Replace invalid characters
            if len(sheet_name) > 31:
                sheet_name = sheet_name[:31]
            
            # Try multiple approaches for maximum compatibility
            output = None
            success = False
            methods_tried = []
            
            # Method 1: Try openpyxl with manual cell writing (most compatible)
            try:
                output = ExcelExporter._create_with_openpyxl_manual(df, sheet_name, format_options)
                success = True
                logger.info("Excel file created using openpyxl manual method")
            except Exception as e:
                methods_tried.append(f"openpyxl_manual: {str(e)}")
                logger.warning(f"openpyxl manual method failed: {str(e)}")
            
            # Method 2: Try pandas ExcelWriter with openpyxl engine (fallback)
            if not success:
                try:
                    output = ExcelExporter._create_with_pandas_openpyxl(df, sheet_name, format_options)
                    success = True
                    logger.info("Excel file created using pandas with openpyxl engine")
                except Exception as e:
                    methods_tried.append(f"pandas_openpyxl: {str(e)}")
                    logger.warning(f"pandas openpyxl method failed: {str(e)}")
            
            # Method 3: Try xlsxwriter engine (optional for deployment platform)
            if not success:
                try:
                    output = ExcelExporter._create_with_xlsxwriter(df, sheet_name, format_options)
                    success = True
                    logger.info("Excel file created using xlsxwriter engine")
                except Exception as e:
                    methods_tried.append(f"xlsxwriter: {str(e)}")
                    logger.warning(f"xlsxwriter method failed: {str(e)}")
            
            # Method 4: Ultimate fallback - basic pandas without formatting
            if not success:
                try:
                    logger.info("Attempting ultimate fallback: basic pandas Excel creation")
                    output = BytesIO()
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
                    output.seek(0)
                    success = True
                    logger.info("Excel file created using basic pandas fallback")
                except Exception as e:
                    methods_tried.append(f"basic_pandas: {str(e)}")
                    logger.error(f"Basic pandas fallback also failed: {str(e)}")
            
            if not success or output is None:
                error_details = "; ".join(methods_tried)
                raise Exception(f"All Excel creation methods failed: {error_details}")
            
            # Validate the created file
            file_content = output.getvalue()
            if len(file_content) < 100:
                raise Exception(f"Generated Excel file is too small ({len(file_content)} bytes), likely corrupted")
            
            if not file_content.startswith(b'PK'):
                raise Exception("Generated file doesn't have valid Excel header (missing PK signature)")
            
            # Try to validate by opening with openpyxl
            try:
                from openpyxl import load_workbook
                output.seek(0)
                test_wb = load_workbook(output)
                test_wb.close()
                output.seek(0)
                logger.debug("Excel file validation passed")
            except Exception as validation_error:
                logger.warning(f"Excel file validation failed, but continuing: {str(validation_error)}")
            
            logger.info(f"Excel file created successfully with {len(df)} rows, size: {len(file_content)} bytes")
            return output
            
        except Exception as e:
            logger.error(f"Excel export failed: {str(e)}")
            raise Exception(f"Excel export failed: {str(e)}")
    
    @staticmethod
    def _create_with_openpyxl_manual(df, sheet_name, format_options):
        """Create Excel file using openpyxl with manual cell writing for maximum compatibility"""
        output = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # Write headers
        for col_num, header in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_num)
            # Ensure header is string and clean
            header_value = ExcelExporter.sanitize_cell_value(str(header))
            cell.value = header_value
        
        # Write data rows
        for row_num, (_, row_data) in enumerate(df.iterrows(), 2):
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                # Clean and convert value
                if pd.isna(value) or value is None:
                    cell.value = ""
                else:
                    # Convert to string first, then sanitize
                    str_value = str(value)
                    clean_value = ExcelExporter.sanitize_cell_value(str_value)
                    
                    # Try to preserve numeric types if possible
                    if str_value.replace('.', '').replace('-', '').isdigit():
                        try:
                            if '.' in str_value:
                                cell.value = float(clean_value)
                            else:
                                cell.value = int(clean_value)
                        except (ValueError, TypeError):
                            cell.value = clean_value
                    else:
                        cell.value = clean_value
        
        # Apply formatting
        if format_options:
            ExcelExporter._apply_formatting(ws, df, format_options)
        else:
            ExcelExporter._apply_default_formatting(ws, df)
        
        # Save with explicit options for compatibility
        wb.save(output)
        output.seek(0)
        return output
    
    @staticmethod
    def _create_with_pandas_openpyxl(df, sheet_name, format_options):
        """Create Excel file using pandas with openpyxl engine"""
        output = BytesIO()
        
        try:
            # Use pandas ExcelWriter with specific options for compatibility
            # Note: options parameter may not be available in all pandas versions
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # Set options on the writer if supported
                try:
                    writer.options = {
                        'remove_timezone': True,
                        'strings_to_formulas': False,
                        'strings_to_urls': False
                    }
                except (AttributeError, TypeError):
                    # Options not supported in this pandas version
                    logger.debug("ExcelWriter options not supported, using defaults")
                
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Apply basic formatting
                workbook = writer.book
                worksheet = writer.sheets[sheet_name]
                ExcelExporter._apply_default_formatting(worksheet, df)
                
        except Exception as e:
            logger.warning(f"pandas openpyxl with options failed, trying basic approach: {str(e)}")
            # Fallback to basic pandas without options
            output = BytesIO()
            df.to_excel(output, sheet_name=sheet_name, index=False, engine='openpyxl')
        
        output.seek(0)
        return output
    
    @staticmethod
    def _create_with_xlsxwriter(df, sheet_name, format_options):
        """Create Excel file using xlsxwriter engine (fallback)"""
        output = BytesIO()
        
        try:
            # Try to use xlsxwriter if available
            try:
                with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                    # Set options if supported
                    try:
                        writer.options = {
                            'remove_timezone': True,
                            'strings_to_formulas': False,
                            'strings_to_urls': False
                        }
                    except (AttributeError, TypeError):
                        logger.debug("ExcelWriter options not supported for xlsxwriter")
                    
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    # Basic formatting with xlsxwriter
                    workbook = writer.book
                    worksheet = writer.sheets[sheet_name]
                    
                    # Header format
                    header_format = workbook.add_format({
                        'bold': True,
                        'bg_color': '#366092',
                        'font_color': 'white',
                        'align': 'center',
                        'valign': 'vcenter'
                    })
                    
                    # Apply header formatting
                    for col_num, column in enumerate(df.columns):
                        worksheet.write(0, col_num, column, header_format)
                    
                    # Auto-adjust column widths
                    for i, column in enumerate(df.columns):
                        max_len = max(
                            df[column].astype(str).apply(len).max(),
                            len(str(column))
                        )
                        worksheet.set_column(i, i, min(max_len + 2, 50))
            except Exception as writer_error:
                # Fallback to basic xlsxwriter without options
                logger.debug(f"xlsxwriter with options failed: {writer_error}")
                output = BytesIO()
                df.to_excel(output, sheet_name=sheet_name, index=False, engine='xlsxwriter')
        
        except ImportError:
            # xlsxwriter not available - this is OK for deployment platform deployment
            logger.info("xlsxwriter not available, will use openpyxl fallback")
            raise Exception("xlsxwriter not available")
        except Exception as e:
            logger.warning(f"xlsxwriter method failed: {str(e)}")
            raise Exception(f"xlsxwriter failed: {str(e)}")
        
        output.seek(0)
        return output
    
    @staticmethod
    def _apply_default_formatting(worksheet, dataframe):
        """Apply default formatting to Excel worksheet with error handling"""
        try:
            # Header formatting
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Apply header formatting
            for col_num, column in enumerate(dataframe.columns, 1):
                try:
                    cell = worksheet.cell(row=1, column=col_num)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                except Exception as e:
                    logger.warning(f"Failed to format header cell {col_num}: {str(e)}")
            
            # Auto-adjust column widths with error handling
            ExcelExporter._adjust_column_widths_safe(worksheet)
            
        except Exception as e:
            logger.warning(f"Failed to apply default formatting: {str(e)}")
    
    @staticmethod
    def _adjust_column_widths_safe(worksheet):
        """Auto-adjust column widths with comprehensive error handling"""
        try:
            for column_cells in worksheet.columns:
                max_length = 0
                column_letter = None
                
                for cell in column_cells:
                    try:
                        if cell.value is not None:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                        
                        # Get column letter from first cell
                        if column_letter is None:
                            column_letter = cell.column_letter
                            
                    except Exception as e:
                        logger.debug(f"Error processing cell in column width adjustment: {str(e)}")
                        continue
                
                # Set column width with reasonable limits
                if column_letter and max_length > 0:
                    try:
                        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                        adjusted_width = max(adjusted_width, 8)   # Minimum width of 8
                        worksheet.column_dimensions[column_letter].width = adjusted_width
                    except Exception as e:
                        logger.debug(f"Failed to set width for column {column_letter}: {str(e)}")
                        
        except Exception as e:
            logger.warning(f"Column width adjustment failed: {str(e)}")
    
    @staticmethod
    def _apply_formatting(worksheet, dataframe, format_options):
        """Apply custom formatting based on format_options with error handling"""
        try:
            # Header formatting
            if format_options.get('header_style'):
                header_style = format_options['header_style']
                
                try:
                    header_font = Font(
                        bold=header_style.get('bold', True),
                        color=header_style.get('font_color', "FFFFFF")
                    )
                    header_fill = PatternFill(
                        start_color=header_style.get('bg_color', "366092"),
                        end_color=header_style.get('bg_color', "366092"),
                        fill_type="solid"
                    )
                    
                    for col_num in range(1, len(dataframe.columns) + 1):
                        try:
                            cell = worksheet.cell(row=1, column=col_num)
                            cell.font = header_font
                            cell.fill = header_fill
                        except Exception as e:
                            logger.debug(f"Failed to format header cell {col_num}: {str(e)}")
                            
                except Exception as e:
                    logger.warning(f"Failed to create header formatting styles: {str(e)}")
            
            # Column width adjustment
            if format_options.get('auto_width', True):
                ExcelExporter._adjust_column_widths_safe(worksheet)
            
            # Custom column widths
            if format_options.get('column_widths'):
                try:
                    for col, width in format_options['column_widths'].items():
                        if col in dataframe.columns:
                            col_index = list(dataframe.columns).index(col) + 1
                            column_letter = worksheet.cell(row=1, column=col_index).column_letter
                            worksheet.column_dimensions[column_letter].width = width
                except Exception as e:
                    logger.warning(f"Failed to apply custom column widths: {str(e)}")
                    
        except Exception as e:
            logger.warning(f"Custom formatting failed: {str(e)}")
    
    @staticmethod
    def _adjust_column_widths(worksheet):
        """Auto-adjust column widths - kept for backward compatibility"""
        ExcelExporter._adjust_column_widths_safe(worksheet)
