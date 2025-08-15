"""
YBB Data Management Service - Railway Production App
Fixed version without Unicode characters that may cause encoding issues
"""
from flask import Flask, request, jsonify, send_file, g
from flask_cors import CORS
import json
import os
from datetime import datetime, timedelta
import tempfile
from io import BytesIO, StringIO
import logging
import uuid
import time
import sys
from logging.handlers import RotatingFileHandler

# Try to import pandas/numpy with fallback
try:
    import pandas as pd
    import numpy as np
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    PANDAS_AVAILABLE = True
    EXCEL_AVAILABLE = True
    print("SUCCESS: Pandas and Excel libraries loaded successfully")
except ImportError as e:
    print(f"WARNING: Advanced libraries not available: {str(e)}")
    print("INFO: Using CSV fallback mode")
    PANDAS_AVAILABLE = False
    EXCEL_AVAILABLE = False

# Import configurations and services with fallback
try:
    from config.ybb_export_config import (
        EXPORT_TEMPLATES, STATUS_MAPPINGS, SYSTEM_CONFIG, 
        get_template, get_status_label, get_chunk_size, should_use_chunked_processing,
        get_cleanup_config, get_storage_limits
    )
    from services.ybb_export_service import YBBExportService
    from utils.file_manager import ExportFileManager
    FULL_SERVICES_AVAILABLE = True
    print("SUCCESS: Full service modules loaded successfully")
except ImportError as e:
    print(f"WARNING: Service modules not available: {str(e)}")
    print("INFO: Using basic implementation")
    FULL_SERVICES_AVAILABLE = False

# Try to import certificate service
try:
    from services.certificate_service import CertificateService
    certificate_service = CertificateService()
    CERTIFICATE_SERVICE_AVAILABLE = True
    print("SUCCESS: Certificate service loaded successfully")
except ImportError as e:
    print(f"WARNING: Certificate service not available: {str(e)}")
    print("INFO: Certificate endpoints will return service unavailable")
    CERTIFICATE_SERVICE_AVAILABLE = False
except Exception as e:
    print(f"WARNING: Certificate service initialization failed: {str(e)}")
    CERTIFICATE_SERVICE_AVAILABLE = False

# Flask app initialization
app = Flask(__name__)
CORS(app, origins="*")

# Configure logging without Unicode characters
def setup_logging():
    if not os.path.exists('logs'):
        os.makedirs('logs')
    
    # Create formatter without Unicode
    formatter = logging.Formatter(
        '%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    )
    
    # File handler
    file_handler = RotatingFileHandler(
        'logs/ybb_api.log', maxBytes=10240000, backupCount=10
    )
    file_handler.setFormatter(formatter)
    file_handler.setLevel(logging.INFO)
    
    # Console handler
    console_handler = logging.StreamHandler(sys.stdout)
    console_handler.setFormatter(formatter)
    console_handler.setLevel(logging.INFO)
    
    # Configure app logger
    app.logger.addHandler(file_handler)
    app.logger.addHandler(console_handler)
    app.logger.setLevel(logging.INFO)
    
    return app.logger

logger = setup_logging()

# Initialize services
if FULL_SERVICES_AVAILABLE:
    try:
        export_service = YBBExportService()
        logger.info("SUCCESS: Full YBB Export Service initialized")
    except Exception as e:
        logger.warning(f"WARNING: Full service failed, using fallback: {str(e)}")
        FULL_SERVICES_AVAILABLE = False

# Fallback storage for basic functionality
if not FULL_SERVICES_AVAILABLE:
    app.exports_storage = {}

@app.before_request
def before_request():
    """Before request handler"""
    g.request_id = str(uuid.uuid4())[:8]
    g.start_time = time.time()
    
    logger.info(
        f"REQUEST_START | ID: {g.request_id} | "
        f"Method: {request.method} | Path: {request.path}"
    )

@app.after_request
def after_request(response):
    """After request handler"""
    duration = round((time.time() - g.start_time) * 1000, 2)
    
    logger.info(
        f"REQUEST_END | ID: {g.request_id} | "
        f"Status: {response.status_code} | Duration: {duration}ms"
    )
    
    return response

@app.route('/health', methods=['GET'])
def health():
    """Comprehensive health check endpoint"""
    return jsonify({
        "status": "healthy",
        "timestamp": datetime.now().isoformat(),
        "service": "YBB Data Processing Service",
        "version": "1.0.0",
        "capabilities": {
            "pandas_available": PANDAS_AVAILABLE,
            "excel_available": EXCEL_AVAILABLE,
            "full_services": FULL_SERVICES_AVAILABLE,
            "certificate_service": CERTIFICATE_SERVICE_AVAILABLE
        },
        "request_id": getattr(g, 'request_id', 'unknown')
    })

def create_csv_export(data, filename=None):
    """Create CSV export using native Python"""
    if not data:
        return None, None
    
    # Create CSV content
    if isinstance(data, list) and len(data) > 0:
        # Get headers from first item
        headers = list(data[0].keys()) if isinstance(data[0], dict) else ["data"]
        
        csv_lines = []
        csv_lines.append(",".join(f'"{header}"' for header in headers))
        
        for item in data:
            if isinstance(item, dict):
                row = []
                for header in headers:
                    value = str(item.get(header, "")).replace('"', '""')
                    row.append(f'"{value}"')
                csv_lines.append(",".join(row))
            else:
                csv_lines.append(f'"' + str(item).replace('"', '""') + '"')
        
        csv_content = "\n".join(csv_lines)
        final_filename = filename or f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        
        return csv_content.encode('utf-8'), final_filename
    
    return None, None

def create_excel_export_fallback(data, filename=None):
    """Create basic Excel export without pandas"""
    if not EXCEL_AVAILABLE or not data:
        return create_csv_export(data, filename)
    
    try:
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Export Data"
        
        # Add headers
        if isinstance(data, list) and len(data) > 0 and isinstance(data[0], dict):
            headers = list(data[0].keys())
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Add data
            for row, item in enumerate(data, 2):
                for col, header in enumerate(headers, 1):
                    ws.cell(row=row, column=col, value=item.get(header, ""))
        
        # Save to bytes
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        final_filename = filename or f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return excel_buffer.read(), final_filename
        
    except Exception as e:
        logger.warning(f"Excel export failed, using CSV: {str(e)}")
        return create_csv_export(data, filename.replace('.xlsx', '.csv') if filename else None)

@app.route('/api/ybb/export/participants', methods=['POST'])
def export_participants():
    """Export participants data with full API compatibility"""
    try:
        request_data = request.get_json()
        
        if not request_data or 'data' not in request_data:
            return jsonify({
                "status": "error",
                "message": "Invalid request: missing 'data' field",
                "request_id": g.request_id
            }), 400
        
        # Use full service if available
        if FULL_SERVICES_AVAILABLE:
            try:
                result = export_service.create_export({
                    "export_type": "participants",
                    "data": request_data['data'],
                    "template": request_data.get('template', 'standard'),
                    "format": request_data.get('format', 'excel'),
                    "filename": request_data.get('filename'),
                    "sheet_name": request_data.get('sheet_name')
                })
                return jsonify(result)
            except Exception as e:
                logger.warning(f"Full service failed, using fallback: {str(e)}")
        
        # Fallback implementation
        export_id = str(uuid.uuid4())
        data = request_data['data']
        filename = request_data.get('filename')
        format_type = request_data.get('format', 'excel').lower()
        
        # Create export content
        if format_type == 'excel':
            file_content, final_filename = create_excel_export_fallback(data, filename)
        else:
            file_content, final_filename = create_csv_export(data, filename)
        
        if file_content is None:
            return jsonify({
                "status": "error",
                "message": "Failed to create export",
                "request_id": g.request_id
            }), 500
        
        # Store export
        app.exports_storage[export_id] = {
            "export_id": export_id,
            "status": "success",
            "export_type": "participants",
            "template": request_data.get('template', 'standard'),
            "format": format_type,
            "record_count": len(data),
            "file_content": file_content,
            "filename": final_filename,
            "created_at": datetime.now(),
            "expires_at": datetime.now() + timedelta(hours=24)
        }
        
        return jsonify({
            "status": "success",
            "message": "Export completed successfully",
            "data": {
                "export_id": export_id,
                "file_name": final_filename,
                "file_size": len(file_content),
                "record_count": len(data),
                "download_url": f"/api/ybb/export/{export_id}/download",
                "expires_at": app.exports_storage[export_id]["expires_at"].isoformat()
            },
            "system_info": {
                "export_type": "participants",
                "template": request_data.get('template', 'standard'),
                "format": format_type,
                "processing_mode": "full_service" if FULL_SERVICES_AVAILABLE else "fallback"
            },
            "request_id": g.request_id
        })
        
    except Exception as e:
        logger.error(f"Export failed: {str(e)}")
        return jsonify({
            "status": "error",
            "message": f"Export failed: {str(e)}",
            "request_id": g.request_id
        }), 500

@app.route('/api/ybb/export/payments', methods=['POST'])
def export_payments():
    """Export payments data"""
    try:
        request_data = request.get_json()
        if request_data:
            request_data['export_type'] = 'payments'
        
        # Forward to participants export with different type
        return export_participants()
        
    except Exception as e:
        logger.error(f"Payments export failed: {str(e)}")
        return jsonify({
            "status": "error",
            "message": f"Payments export failed: {str(e)}",
            "request_id": getattr(g, 'request_id', 'unknown')
        }), 500

@app.route('/api/ybb/export/ambassadors', methods=['POST'])
def export_ambassadors():
    """Export ambassadors data"""
    try:
        request_data = request.get_json()
        if request_data:
            request_data['export_type'] = 'ambassadors'
        
        # Forward to participants export with different type
        return export_participants()
        
    except Exception as e:
        logger.error(f"Ambassadors export failed: {str(e)}")
        return jsonify({
            "status": "error",
            "message": f"Ambassadors export failed: {str(e)}",
            "request_id": getattr(g, 'request_id', 'unknown')
        }), 500

@app.route('/api/ybb/export/<export_id>/download', methods=['GET'])
def download_export(export_id):
    """Download export file"""
    try:
        # Use full service if available
        if FULL_SERVICES_AVAILABLE and hasattr(export_service, 'download_export'):
            try:
                file_content, filename = export_service.download_export(export_id)
                if file_content:
                    # Determine content type
                    if filename.endswith('.xlsx'):
                        content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                    elif filename.endswith('.csv'):
                        content_type = 'text/csv'
                    else:
                        content_type = 'application/octet-stream'
                    
                    return send_file(
                        BytesIO(file_content),
                        as_attachment=True,
                        download_name=filename,
                        mimetype=content_type
                    )
            except Exception as e:
                logger.warning(f"Full service download failed, using fallback: {str(e)}")
        
        # Fallback implementation
        if export_id not in app.exports_storage:
            return jsonify({
                "status": "error",
                "message": "Export file not found or expired",
                "request_id": getattr(g, 'request_id', 'unknown')
            }), 404
        
        export_info = app.exports_storage[export_id]
        
        # Check expiration
        if export_info["expires_at"] < datetime.now():
            del app.exports_storage[export_id]
            return jsonify({
                "status": "error",
                "message": "Export file has expired",
                "request_id": getattr(g, 'request_id', 'unknown')
            }), 404
        
        # Determine content type
        filename = export_info["filename"]
        if filename.endswith('.xlsx'):
            content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        elif filename.endswith('.csv'):
            content_type = 'text/csv'
        else:
            content_type = 'application/octet-stream'
        
        return send_file(
            BytesIO(export_info["file_content"]),
            as_attachment=True,
            download_name=filename,
            mimetype=content_type
        )
        
    except Exception as e:
        logger.error(f"Download failed: {str(e)}")
        return jsonify({
            "status": "error",
            "message": f"Download failed: {str(e)}",
            "request_id": getattr(g, 'request_id', 'unknown')
        }), 500

@app.route('/api/ybb/export/<export_id>/status', methods=['GET'])
def export_status(export_id):
    """Get export status"""
    try:
        # Use full service if available
        if FULL_SERVICES_AVAILABLE and hasattr(export_service, 'get_export_status'):
            try:
                status = export_service.get_export_status(export_id)
                if status:
                    return jsonify(status)
            except Exception as e:
                logger.warning(f"Full service status check failed: {str(e)}")
        
        # Fallback implementation
        if export_id not in app.exports_storage:
            return jsonify({
                "status": "error",
                "message": "Export not found",
                "request_id": getattr(g, 'request_id', 'unknown')
            }), 404
        
        export_info = app.exports_storage[export_id]
        
        return jsonify({
            "status": "success",
            "data": {
                "export_id": export_id,
                "status": export_info.get("status", "completed"),
                "export_type": export_info.get("export_type", "unknown"),
                "record_count": export_info.get("record_count", 0),
                "file_size": len(export_info.get("file_content", b"")),
                "created_at": export_info.get("created_at", datetime.now()).isoformat(),
                "expires_at": export_info.get("expires_at", datetime.now()).isoformat()
            },
            "request_id": getattr(g, 'request_id', 'unknown')
        })
        
    except Exception as e:
        logger.error(f"Status check failed: {str(e)}")
        return jsonify({
            "status": "error",
            "message": f"Status check failed: {str(e)}",
            "request_id": getattr(g, 'request_id', 'unknown')
        }), 500

# Error handlers
@app.errorhandler(400)
def bad_request(error):
    return jsonify({
        "status": "error",
        "message": "Bad request - invalid data format",
        "request_id": getattr(g, 'request_id', 'unknown')
    }), 400

@app.errorhandler(404)
def not_found(error):
    return jsonify({
        "error": "Endpoint not found",
        "method": request.method,
        "url": request.url,
        "request_id": getattr(g, 'request_id', 'unknown')
    }), 404

@app.errorhandler(500)
def internal_error(error):
    return jsonify({
        "status": "error",
        "message": "Internal server error",
        "request_id": getattr(g, 'request_id', 'unknown')
    }), 500

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    host = os.environ.get('HOST', '0.0.0.0')
    
    logger.info(f"Starting YBB Data Management Service")
    logger.info(f"Host: {host}:{port}")
    logger.info(f"Pandas: {'Available' if PANDAS_AVAILABLE else 'Fallback'}")
    logger.info(f"Excel: {'Available' if EXCEL_AVAILABLE else 'CSV Only'}")
    logger.info(f"Services: {'Full' if FULL_SERVICES_AVAILABLE else 'Basic'}")
    logger.info(f"Certificates: {'Available' if CERTIFICATE_SERVICE_AVAILABLE else 'Unavailable'}")
    
    app.run(host=host, port=port, debug=False)
