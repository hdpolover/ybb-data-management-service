"""
Test Excel file creation exactly as the service does it
"""
import sys
import os
sys.path.append('.')

from services.ybb_export_service import YBBExportService
import tempfile
import json

def test_service_excel_generation():
    """Test Excel generation exactly as the service does it"""
    
    # Initialize service
    export_service = YBBExportService()
    
    # Sample data similar to what you might have for participants
    test_data = []
    for i in range(10):
        test_data.append({
            "id": i + 1,
            "full_name": f"Test User {i + 1}",
            "email": f"user{i + 1}@example.com",
            "nationality": "Test Country",
            "institution": "Test University",
            "phone_number": f"+1234567890{i}",
            "category": "test",
            "form_status": "approved",
            "payment_status": "paid",
            "created_at": "2025-08-14T10:00:00"
        })
    
    print("Testing service Excel generation...\n")
    
    # Test 1: Standard single file export
    try:
        print("Test 1: Standard single file export")
        export_request = {
            "export_type": "participants",
            "data": test_data,
            "template": "standard",
            "format": "excel",
            "filename": "test_participants.xlsx"
        }
        
        result = export_service.create_export(export_request)
        
        if result["status"] == "success":
            export_id = result["data"]["export_id"]
            print(f"✅ Export created successfully: {export_id}")
            print(f"Strategy: {result['export_strategy']}")
            print(f"Filename: {result['data']['file_name']}")
            
            # Try to get the file content
            export_info = export_service.exports_storage.get(export_id)
            if export_info and "file_content" in export_info:
                file_content = export_info["file_content"]
                print(f"File size: {len(file_content)} bytes")
                print(f"Header check: {file_content[:4]} (should be b'PK\\x03\\x04')")
                
                # Save to temp file and validate
                temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
                temp_file.write(file_content)
                temp_file.close()
                
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(temp_file.name)
                    print(f"✅ Excel validation passed - {len(wb.sheetnames)} sheets")
                    print(f"Sheet names: {wb.sheetnames}")
                    
                    # Check data
                    ws = wb.active
                    print(f"Rows: {ws.max_row}, Columns: {ws.max_column}")
                    wb.close()
                    
                except Exception as ve:
                    print(f"❌ Excel validation failed: {ve}")
                
                os.unlink(temp_file.name)
            else:
                print("❌ No file content found in export info")
        else:
            print(f"❌ Export failed: {result}")
        
        print()
        
    except Exception as e:
        print(f"❌ Test failed: {e}\n")
        import traceback
        traceback.print_exc()
    
    # Test 2: Force chunking for small dataset
    try:
        print("Test 2: Force chunking (small dataset)")
        export_request = {
            "export_type": "participants",
            "data": test_data,
            "template": "complete",
            "format": "excel",
            "force_chunking": True,
            "chunk_size": 3,
            "filename": "test_chunked_participants.xlsx"
        }
        
        result = export_service.create_export(export_request)
        
        if result["status"] == "success":
            export_id = result["data"]["export_id"]
            print(f"✅ Chunked export created successfully: {export_id}")
            print(f"Strategy: {result['export_strategy']}")
            print(f"Total files: {result['data']['total_files']}")
            
            # Try to get the ZIP file content
            export_info = export_service.exports_storage.get(export_id)
            if export_info and "zip_path" in export_info:
                zip_path = export_info["zip_path"]
                if os.path.exists(zip_path):
                    zip_size = os.path.getsize(zip_path)
                    print(f"ZIP file size: {zip_size} bytes")
                    print(f"ZIP path: {zip_path}")
                    
                    # Validate ZIP file
                    import zipfile
                    try:
                        with zipfile.ZipFile(zip_path, 'r') as zf:
                            file_list = zf.namelist()
                            print(f"✅ ZIP validation passed - {len(file_list)} files")
                            print(f"Files in ZIP: {file_list}")
                    except Exception as zve:
                        print(f"❌ ZIP validation failed: {zve}")
                else:
                    print("❌ ZIP file not found at expected path")
            else:
                print("❌ No ZIP path found in export info")
        else:
            print(f"❌ Chunked export failed: {result}")
        
        print()
        
    except Exception as e:
        print(f"❌ Chunked test failed: {e}\n")
        import traceback
        traceback.print_exc()
    
    print("Service testing completed.")

if __name__ == "__main__":
    test_service_excel_generation()
