#!/usr/bin/env python3
"""
Debug Excel file generation issue
"""
import sys
import os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import requests
import tempfile
import pandas as pd
from openpyxl import load_workbook
import binascii

def test_actual_api_data():
    """Test with actual API call and examine the generated file"""
    
    print("🔍 Testing actual API data generation...")
    
    # Use realistic test data similar to what might come from your system
    test_data = {
        "export_type": "participants",
        "template": "standard", 
        "format": "excel",
        "data": [
            {
                "id": 1,
                "full_name": "John Doe",
                "email": "john.doe@example.com",
                "phone": "+1234567890",
                "institution": "University of Example",
                "form_status": "approved",
                "created_at": "2024-01-15 10:30:00",
                "updated_at": "2024-01-15 10:30:00"
            },
            {
                "id": 2,
                "full_name": "Jane Smith",
                "email": "jane.smith@example.com", 
                "phone": "+0987654321",
                "institution": "Example College",
                "form_status": "pending",
                "created_at": "2024-01-16 14:20:00",
                "updated_at": "2024-01-16 14:20:00"
            },
            {
                "id": 3,
                "full_name": "Bob Johnson",
                "email": "bob.johnson@example.com",
                "phone": "+1122334455", 
                "institution": "Test University",
                "form_status": "rejected",
                "created_at": "2024-01-17 09:15:00",
                "updated_at": "2024-01-17 09:15:00"
            }
        ]
    }
    
    try:
        print("Making API request...")
        response = requests.post(
            "http://127.0.0.1:5000/api/ybb/export/participants",
            json=test_data,
            timeout=30
        )
        
        if response.status_code != 200:
            print(f"❌ API request failed: {response.status_code}")
            print(f"Response: {response.text}")
            return False
            
        result = response.json()
        export_id = result.get('data', {}).get('export_id')
        
        if not export_id:
            print("❌ No export ID returned")
            return False
            
        print(f"✅ Export created with ID: {export_id}")
        
        # Download the file
        print("Downloading Excel file...")
        download_response = requests.get(
            f"http://127.0.0.1:5000/api/ybb/export/{export_id}/download",
            timeout=30
        )
        
        if download_response.status_code != 200:
            print(f"❌ Download failed: {download_response.status_code}")
            return False
            
        excel_content = download_response.content
        print(f"✅ Downloaded {len(excel_content)} bytes")
        
        # Examine the file content
        print("\n🔍 File Analysis:")
        print(f"File size: {len(excel_content)} bytes")
        print(f"First 50 bytes (hex): {binascii.hexlify(excel_content[:50]).decode()}")
        print(f"First 20 bytes (ascii): {excel_content[:20]}")
        
        # Check if it starts with proper Excel signature
        if excel_content.startswith(b'PK'):
            print("✅ File has correct ZIP/Excel signature (PK)")
        else:
            print("❌ File does not have correct Excel signature")
            
        # Save to temp file and test opening
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
            temp_file.write(excel_content)
            temp_path = temp_file.name
            
        print(f"Saved to: {temp_path}")
        
        # Test with different libraries
        success_tests = 0
        total_tests = 3
        
        # Test 1: openpyxl
        try:
            workbook = load_workbook(temp_path)
            sheet = workbook.active
            print(f"✅ openpyxl: Successfully opened, {sheet.max_row} rows, {sheet.max_column} columns")
            
            # Read a few cells to check data
            print("Sample data:")
            for row in range(1, min(4, sheet.max_row + 1)):
                row_data = []
                for col in range(1, min(4, sheet.max_column + 1)):
                    cell_value = sheet.cell(row=row, column=col).value
                    row_data.append(str(cell_value)[:20] if cell_value else "")
                print(f"  Row {row}: {row_data}")
                
            success_tests += 1
        except Exception as e:
            print(f"❌ openpyxl failed: {str(e)}")
            
        # Test 2: pandas
        try:
            df = pd.read_excel(temp_path)
            print(f"✅ pandas: Successfully opened, shape {df.shape}")
            print(f"Columns: {list(df.columns)}")
            if not df.empty:
                print("First row data:")
                for col in df.columns:
                    value = df.iloc[0][col] if len(df) > 0 else ""
                    print(f"  {col}: {str(value)[:30]}")
            success_tests += 1
        except Exception as e:
            print(f"❌ pandas failed: {str(e)}")
            
        # Test 3: Try manual inspection
        try:
            from zipfile import ZipFile
            with ZipFile(temp_path, 'r') as zip_file:
                file_list = zip_file.namelist()
                print(f"✅ ZIP inspection: Found {len(file_list)} files in Excel archive")
                print(f"Files: {file_list[:5]}")  # Show first 5 files
                success_tests += 1
        except Exception as e:
            print(f"❌ ZIP inspection failed: {str(e)}")
            
        print(f"\n📊 Test Results: {success_tests}/{total_tests} tests passed")
        
        # Keep the file for manual inspection
        print(f"\n📁 File saved at: {temp_path}")
        print("You can manually try to open this file in Excel to see the exact error message.")
        
        return success_tests >= 2
        
    except requests.exceptions.ConnectionError:
        print("❌ Cannot connect to API server. Make sure it's running on http://127.0.0.1:5000")
        return False
    except Exception as e:
        print(f"❌ Test failed: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

def test_local_excel_generation():
    """Test local Excel generation with problematic data"""
    
    print("\n🧪 Testing local Excel generation...")
    
    # Test with data that might have issues
    problematic_data = [
        {
            "id": 1,
            "name": "Normal Name",
            "email": "normal@test.com",
            "notes": "Regular content"
        },
        {
            "id": 2,
            "name": "Name with émojis 🎉 and ünicøde",
            "email": "unicode@test.com",
            "notes": "Content with special chars: àáâãäåæçèéêë"
        },
        {
            "id": 3,
            "name": "Very long name " + "x" * 100,
            "email": "long@test.com", 
            "notes": "Very long notes " + "y" * 1000
        }
    ]
    
    try:
        from utils.excel_exporter import ExcelExporter
        
        print("Creating Excel file locally...")
        excel_output = ExcelExporter.create_excel_file(
            data=problematic_data,
            filename="debug_test.xlsx",
            sheet_name="Debug Test"
        )
        
        content = excel_output.getvalue()
        print(f"Generated {len(content)} bytes")
        
        # Check file signature
        if content.startswith(b'PK'):
            print("✅ Local file has correct signature")
        else:
            print("❌ Local file has incorrect signature")
            print(f"First 20 bytes: {content[:20]}")
            
        # Save and test
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
            temp_file.write(content)
            temp_path = temp_file.name
            
        # Test opening
        try:
            df = pd.read_excel(temp_path)
            print(f"✅ Local file opens successfully: {df.shape}")
            return True
        except Exception as e:
            print(f"❌ Local file failed to open: {str(e)}")
            return False
            
    except Exception as e:
        print(f"❌ Local generation failed: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    print("🔧 Excel File Debug Analysis")
    print("=" * 50)
    
    # Test local generation first
    local_ok = test_local_excel_generation()
    
    # Test API if local works
    if local_ok:
        api_ok = test_actual_api_data()
    else:
        print("❌ Skipping API test due to local generation failure")
        api_ok = False
        
    print("\n" + "=" * 50)
    print("📊 Debug Results:")
    print(f"Local Generation: {'✅ PASSED' if local_ok else '❌ FAILED'}")
    print(f"API Generation: {'✅ PASSED' if api_ok else '❌ FAILED'}")
    
    if not local_ok or not api_ok:
        print("\n🚨 Issue detected! Check the analysis above for specific problems.")
    else:
        print("\n✅ Both tests passed. The issue might be with specific data or Excel version.")
