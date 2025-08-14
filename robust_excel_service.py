"""
Enhanced Excel Service - Fixed version with comprehensive validation
"""
from utils.excel_exporter import ExcelExporter
from io import BytesIO
import tempfile
import os
from openpyxl import load_workbook, Workbook
import mimetypes
import logging

logger = logging.getLogger(__name__)

class RobustExcelService:
    """Enhanced Excel service with comprehensive validation and fixing"""
    
    @staticmethod
    def create_excel_file_robust(data, filename=None, sheet_name="Data"):
        """
        Create Excel file with multiple fallbacks and validation
        
        Args:
            data: List of dictionaries or pandas DataFrame
            filename: Output filename (will be validated/fixed)
            sheet_name: Excel sheet name
            
        Returns:
            tuple: (file_content_bytes, validated_filename, validation_info)
        """
        
        # Step 1: Fix filename
        fixed_filename = RobustExcelService._fix_filename(filename)
        logger.info(f"Creating Excel file: {fixed_filename}")
        
        # Step 2: Try multiple creation methods
        methods = [
            ("ExcelExporter", RobustExcelService._create_with_excel_exporter),
            ("pandas_basic", RobustExcelService._create_with_pandas_basic),
            ("openpyxl_manual", RobustExcelService._create_with_openpyxl)
        ]
        
        for method_name, method_func in methods:
            try:
                logger.info(f"Attempting Excel creation with: {method_name}")
                
                file_content = method_func(data, sheet_name)
                
                # Validate the created file
                validation_info = RobustExcelService._validate_excel_content(file_content)
                
                if validation_info['valid']:
                    logger.info(f"✅ Excel creation successful with {method_name}: {len(file_content)} bytes")
                    return file_content, fixed_filename, validation_info
                else:
                    logger.warning(f"❌ {method_name} created invalid file: {validation_info}")
                    
            except Exception as e:
                logger.warning(f"❌ {method_name} failed: {str(e)}")
                continue
        
        # If all methods fail, raise an exception
        raise Exception("All Excel creation methods failed. Check data format and dependencies.")
    
    @staticmethod
    def _fix_filename(filename):
        """Fix and validate filename"""
        if not filename:
            from datetime import datetime
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"export_{timestamp}.xlsx"
        
        # Ensure .xlsx extension
        if not filename.lower().endswith('.xlsx'):
            # Remove other extensions and add .xlsx
            base_name = os.path.splitext(filename)[0]
            filename = f"{base_name}.xlsx"
        
        # Sanitize filename for filesystem compatibility
        import re
        filename = re.sub(r'[<>:"/\\|?*]', '_', filename)
        filename = re.sub(r'_+', '_', filename)
        filename = filename.strip()
        
        return filename
    
    @staticmethod
    def _create_with_excel_exporter(data, sheet_name):
        """Create using the existing ExcelExporter"""
        excel_output = ExcelExporter.create_excel_file(
            data=data,
            sheet_name=sheet_name,
            format_options=None
        )
        return excel_output.getvalue()
    
    @staticmethod
    def _create_with_pandas_basic(data, sheet_name):
        """Create using basic pandas (most reliable)"""
        import pandas as pd
        
        # Convert to DataFrame if needed
        if isinstance(data, list):
            df = pd.DataFrame(data)
        else:
            df = data.copy()
        
        # Basic data cleaning
        for col in df.columns:
            df[col] = df[col].astype(str).apply(lambda x: str(x)[:32767] if x else "")
        
        # Create Excel file
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name=sheet_name)
        
        output.seek(0)
        return output.getvalue()
    
    @staticmethod
    def _create_with_openpyxl(data, sheet_name):
        """Create using manual openpyxl (most compatible)"""
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name[:31]  # Excel sheet name limit
        
        # Convert to list of dicts if needed
        if hasattr(data, 'to_dict'):  # pandas DataFrame
            data = data.to_dict('records')
        
        if not data:
            raise ValueError("No data to export")
        
        # Write headers
        headers = list(data[0].keys())
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=str(header))
        
        # Write data
        for row_num, record in enumerate(data, 2):
            for col_num, header in enumerate(headers, 1):
                value = record.get(header, "")
                # Excel cell value limits
                if isinstance(value, str) and len(value) > 32767:
                    value = value[:32764] + "..."
                ws.cell(row=row_num, column=col_num, value=value)
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
    
    @staticmethod
    def _validate_excel_content(content):
        """Comprehensive Excel content validation"""
        validation_info = {
            'valid': False,
            'size_ok': False,
            'header_ok': False,
            'excel_readable': False,
            'sheets_count': 0,
            'file_size': len(content),
            'errors': []
        }
        
        # Size check
        if len(content) > 100:
            validation_info['size_ok'] = True
        else:
            validation_info['errors'].append(f"File too small: {len(content)} bytes")
        
        # Header check
        if content.startswith(b'PK\x03\x04'):
            validation_info['header_ok'] = True
        else:
            validation_info['errors'].append(f"Invalid header: {content[:10]}")
        
        # Excel readability check
        if validation_info['size_ok'] and validation_info['header_ok']:
            try:
                # Create temp file for validation
                temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
                temp_file.write(content)
                temp_file.close()
                
                # Try to open with openpyxl
                wb = load_workbook(temp_file.name)
                validation_info['excel_readable'] = True
                validation_info['sheets_count'] = len(wb.sheetnames)
                validation_info['sheet_names'] = wb.sheetnames
                wb.close()
                
                # Cleanup
                os.unlink(temp_file.name)
                
            except Exception as e:
                validation_info['errors'].append(f"Excel read failed: {str(e)}")
        
        # Overall validity
        validation_info['valid'] = (
            validation_info['size_ok'] and 
            validation_info['header_ok'] and 
            validation_info['excel_readable']
        )
        
        return validation_info
    
    @staticmethod
    def get_content_type_headers(filename):
        """Get proper HTTP headers for Excel file serving"""
        
        # Ensure filename has .xlsx extension
        if not filename.lower().endswith('.xlsx'):
            filename = f"{os.path.splitext(filename)[0]}.xlsx"
        
        headers = {
            'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'Content-Disposition': f'attachment; filename="{filename}"',
            'Cache-Control': 'no-cache, no-store, must-revalidate',
            'Pragma': 'no-cache',
            'Expires': '0'
        }
        
        return headers

# Test the robust service
if __name__ == "__main__":
    print("Testing RobustExcelService...\n")
    
    test_data = [
        {"id": 1, "name": "Test User 1", "email": "test1@example.com"},
        {"id": 2, "name": "Test User 2", "email": "test2@example.com"},
        {"id": 3, "name": "Test User 3", "email": "test3@example.com"}
    ]
    
    try:
        content, filename, validation = RobustExcelService.create_excel_file_robust(
            data=test_data,
            filename="robust_test_file",
            sheet_name="Test Data"
        )
        
        print(f"✅ Success!")
        print(f"Filename: {filename}")
        print(f"File size: {len(content)} bytes")
        print(f"Validation: {validation}")
        
        # Test headers
        headers = RobustExcelService.get_content_type_headers(filename)
        print(f"HTTP Headers: {headers}")
        
        # Save and test
        with open(f"test_{filename}", 'wb') as f:
            f.write(content)
        
        print(f"\n✅ Test file saved as: test_{filename}")
        print("Try opening this file in Excel to verify it works!")
        
    except Exception as e:
        print(f"❌ Test failed: {e}")
        import traceback
        traceback.print_exc()
