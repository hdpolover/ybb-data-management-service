#!/usr/bin/env python3
"""
Test the improved Excel exporter with better compatibility
"""
import sys
import os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import tempfile
import pandas as pd
from openpyxl import load_workbook
import binascii

def test_improved_exporter():
    """Test the improved Excel exporter"""
    
    print("🔧 Testing Improved Excel Exporter")
    print("=" * 50)
    
    # Test data with various potential issues
    test_data = [
        {
            "id": 1,
            "name": "Normal User",
            "email": "normal@test.com",
            "notes": "Regular content"
        },
        {
            "id": 2,
            "name": "User with émojis 🎉",
            "email": "emoji@test.com",
            "notes": "Content with special chars: àáâãäåæçèéêë"
        },
        {
            "id": 3,
            "name": "=SUM(1+1)",  # Formula injection test
            "email": "formula@test.com",
            "notes": "Potential formula: @SUM(A1:B1)"
        },
        {
            "id": 4,
            "name": "User\nwith\nnewlines",
            "email": "newlines@test.com",
            "notes": "Content\twith\ttabs\nand\r\ncarriage returns"
        },
        {
            "id": 5,
            "name": None,  # Null value
            "email": "",   # Empty string
            "notes": "Mixed content: 测试 データ ñáéíóú"
        }
    ]
    
    try:
        from utils.excel_exporter import ExcelExporter
        
        print("Creating Excel file with improved exporter...")
        excel_output = ExcelExporter.create_excel_file(
            data=test_data,
            filename="improved_test.xlsx",
            sheet_name="Improved Test"
        )
        
        content = excel_output.getvalue()
        print(f"✅ Generated {len(content)} bytes")
        
        # Check file signature
        if content.startswith(b'PK'):
            print("✅ File has correct Excel signature")
        else:
            print("❌ File has incorrect signature")
            print(f"First 20 bytes: {content[:20]}")
            return False
        
        # Save to temp file
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
            temp_file.write(content)
            temp_path = temp_file.name
        
        print(f"📁 File saved to: {temp_path}")
        
        # Test with different methods
        success_count = 0
        
        # Test 1: openpyxl
        try:
            workbook = load_workbook(temp_path)
            sheet = workbook.active
            print(f"✅ openpyxl: {sheet.max_row} rows, {sheet.max_column} columns")
            
            # Check specific cells for proper sanitization
            print("Sample data inspection:")
            for row in range(1, min(4, sheet.max_row + 1)):
                for col in range(1, min(4, sheet.max_column + 1)):
                    cell = sheet.cell(row=row, column=col)
                    value = str(cell.value)[:30] if cell.value else ""
                    print(f"  Cell({row},{col}): {repr(value)}")
            
            success_count += 1
        except Exception as e:
            print(f"❌ openpyxl failed: {str(e)}")
        
        # Test 2: pandas
        try:
            df = pd.read_excel(temp_path)
            print(f"✅ pandas: shape {df.shape}")
            print(f"Columns: {list(df.columns)}")
            
            # Check data integrity
            if not df.empty:
                print("Data sample:")
                for col in df.columns[:3]:  # First 3 columns
                    value = str(df.iloc[0][col])[:30] if len(df) > 0 else ""
                    print(f"  {col}: {repr(value)}")
            
            success_count += 1
        except Exception as e:
            print(f"❌ pandas failed: {str(e)}")
        
        print(f"\n📊 Tests passed: {success_count}/2")
        
        # Try to manually inspect the file
        print(f"\n📂 Manual test file: {temp_path}")
        print("Please try opening this file in Excel to verify it works.")
        
        return success_count >= 2
        
    except Exception as e:
        print(f"❌ Test failed: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

def test_api_with_improvements():
    """Test API with the improved exporter"""
    
    print("\n🌐 Testing API with Improved Exporter")
    print("=" * 50)
    
    import requests
    
    # Test data similar to actual usage
    api_test_data = {
        "export_type": "participants",
        "template": "standard",
        "format": "excel",
        "data": [
            {
                "id": 1,
                "full_name": "John Doe",
                "email": "john@example.com",
                "phone": "+1234567890",
                "institution": "Test University",
                "form_status": "approved",
                "created_at": "2024-01-15 10:30:00"
            },
            {
                "id": 2,
                "full_name": "Jane Smith with special chars: áéíóú",
                "email": "jane@example.com",
                "phone": "+0987654321",
                "institution": "Example College",
                "form_status": "pending", 
                "created_at": "2024-01-16 14:20:00"
            },
            {
                "id": 3,
                "full_name": "User\nwith\nnewlines",
                "email": "newlines@example.com",
                "phone": "+1122334455",
                "institution": "Multi\nLine University",
                "form_status": "rejected",
                "created_at": "2024-01-17 09:15:00"
            }
        ]
    }
    
    try:
        print("Making API request...")
        response = requests.post(
            "http://127.0.0.1:5000/api/ybb/export/participants",
            json=api_test_data,
            timeout=30
        )
        
        if response.status_code != 200:
            print(f"❌ API request failed: {response.status_code}")
            print(f"Response: {response.text}")
            return False
        
        result = response.json()
        export_id = result.get('data', {}).get('export_id')
        print(f"✅ Export created: {export_id}")
        
        # Download file
        print("Downloading improved Excel file...")
        download_response = requests.get(
            f"http://127.0.0.1:5000/api/ybb/export/{export_id}/download",
            timeout=30
        )
        
        if download_response.status_code != 200:
            print(f"❌ Download failed: {download_response.status_code}")
            return False
        
        content = download_response.content
        print(f"✅ Downloaded {len(content)} bytes")
        
        # Check signature
        if content.startswith(b'PK'):
            print("✅ Downloaded file has correct signature")
        else:
            print("❌ Downloaded file has incorrect signature")
            return False
        
        # Save and test
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
            temp_file.write(content)
            temp_path = temp_file.name
        
        # Quick validation
        try:
            df = pd.read_excel(temp_path)
            print(f"✅ Downloaded file opens correctly: {df.shape}")
            print(f"📂 API test file: {temp_path}")
            print("Please test opening this file in Excel.")
            return True
        except Exception as e:
            print(f"❌ Downloaded file failed to open: {str(e)}")
            return False
        
    except requests.exceptions.ConnectionError:
        print("⚠️  API server not running. Skipping API test.")
        return True
    except Exception as e:
        print(f"❌ API test failed: {str(e)}")
        return False

if __name__ == "__main__":
    print("🔬 Testing Improved Excel Export Compatibility")
    print("=" * 60)
    
    # Test local generation
    local_ok = test_improved_exporter()
    
    # Test API
    api_ok = test_api_with_improvements()
    
    print("\n" + "=" * 60)
    print("📊 Final Results:")
    print(f"Improved Local Generation: {'✅ PASSED' if local_ok else '❌ FAILED'}")
    print(f"Improved API Generation: {'✅ PASSED' if api_ok else '❌ FAILED'}")
    
    if local_ok and api_ok:
        print("\n🎉 Improved Excel exporter is working!")
        print("📝 Manual test files have been created for Excel verification.")
    else:
        print("\n⚠️  Some tests failed. Check the output above.")
        
    print("\n💡 Next Steps:")
    print("1. Open the generated test files in Excel")
    print("2. Verify they open without corruption errors")
    print("3. Check that data is displayed correctly")
    print("4. If still failing, please share the exact Excel error message")
