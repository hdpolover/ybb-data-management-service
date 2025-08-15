#!/usr/bin/env python3
"""
Excel Compatibility Diagnostic Tool
Tests various potential corruption causes
"""
import sys
import os
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from services.ybb_export_service import YBBExportService
from utils.excel_exporter import ExcelExporter
import pandas as pd
from io import BytesIO
import logging

# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def test_excel_compatibility():
    """Comprehensive Excel compatibility test"""
    
    print("🔧 Excel Compatibility Diagnostic Tool")
    print("=" * 50)
    
    test_cases = [
        {
            "name": "Basic Data",
            "data": [{"id": 1, "name": "Test", "email": "test@test.com"}]
        },
        {
            "name": "Long Filename Data", 
            "data": [{"id": 1, "full_name": "Japan_Youth_Summit_2025_Participants_Participants_Approved_Forms_27-07-2025"}]
        },
        {
            "name": "Special Characters",
            "data": [{"id": 1, "name": "José María", "notes": "Special: àáâãäåæçèéêë"}]
        },
        {
            "name": "Unicode Characters",
            "data": [{"id": 1, "name": "Test", "notes": "Unicode: 你好世界 🌍🎌"}]
        },
        {
            "name": "Control Characters",
            "data": [{"id": 1, "name": "Test\nName", "notes": "Tab:\tControl\x00Char"}]
        },
        {
            "name": "Formula Injection",
            "data": [{"id": 1, "name": "=SUM(A1:B1)", "formula": "@SUM(1+1)"}]
        },
        {
            "name": "Large Text",
            "data": [{"id": 1, "description": "A" * 35000}]  # Exceeds Excel cell limit
        },
        {
            "name": "Empty/Null Values",
            "data": [{"id": 1, "name": None, "empty": "", "null": None}]
        }
    ]
    
    success_count = 0
    total_tests = len(test_cases)
    
    for i, test_case in enumerate(test_cases, 1):
        print(f"\n📋 Test {i}/{total_tests}: {test_case['name']}")
        print("-" * 30)
        
        try:
            # Test 1: Direct Excel creation
            print("  Testing direct Excel creation...")
            output = ExcelExporter.create_excel_file(
                data=test_case['data'],
                sheet_name="Test Sheet"
            )
            
            file_content = output.getvalue()
            test_filename = f"test_{i}_{test_case['name'].replace(' ', '_').lower()}.xlsx"
            
            # Write file
            with open(test_filename, 'wb') as f:
                f.write(file_content)
            
            # Validate file
            from openpyxl import load_workbook
            wb = load_workbook(test_filename)
            ws = wb.active
            
            print(f"  ✅ File created: {len(file_content)} bytes")
            print(f"  ✅ File validated: {ws.max_row} rows, {ws.max_column} cols")
            
            # Check first few cells
            if ws.max_row > 1:
                sample_data = []
                for col in range(1, min(4, ws.max_column + 1)):
                    cell_value = ws.cell(row=2, column=col).value
                    if cell_value and len(str(cell_value)) > 20:
                        sample_data.append(f"{str(cell_value)[:20]}...")
                    else:
                        sample_data.append(str(cell_value))
                print(f"  📊 Sample data: {sample_data}")
            
            wb.close()
            
            # Test 2: Via YBB Export Service
            print("  Testing via YBB Export Service...")
            export_request = {
                'export_type': 'participants',
                'template': 'standard',
                'data': test_case['data'],
                'format': 'excel'
            }
            
            export_service = YBBExportService()
            result = export_service.create_export(export_request)
            
            if result['status'] == 'success':
                export_id = result['data']['export_id']
                service_content, service_filename = export_service.download_export(export_id, 'single')
                
                service_test_filename = f"service_{test_filename}"
                with open(service_test_filename, 'wb') as f:
                    f.write(service_content)
                
                # Validate service-created file
                wb2 = load_workbook(service_test_filename)
                print(f"  ✅ Service file: {len(service_content)} bytes")
                print(f"  ✅ Service validated: {wb2.active.max_row} rows")
                wb2.close()
            else:
                print(f"  ❌ Service creation failed: {result.get('message')}")
            
            success_count += 1
            print(f"  🎉 Test PASSED")
            
        except Exception as e:
            print(f"  ❌ Test FAILED: {str(e)}")
            
            # Additional diagnostics for failed tests
            try:
                if 'file_content' in locals():
                    print(f"  🔍 File content analysis:")
                    print(f"     - Size: {len(file_content)} bytes")
                    print(f"     - Header: {file_content[:20].hex()}")
                    print(f"     - PK signature: {file_content.startswith(b'PK')}")
            except:
                pass
    
    print(f"\n📊 Final Results")
    print("=" * 50)
    print(f"✅ Passed: {success_count}/{total_tests}")
    print(f"❌ Failed: {total_tests - success_count}/{total_tests}")
    
    if success_count == total_tests:
        print("🎉 All tests passed! Excel generation is working correctly.")
        print("   The issue may be environment-specific or related to Excel version.")
        print("\n🔧 Recommendations:")
        print("   1. Try opening the file with different Excel versions")
        print("   2. Check if your Excel has security restrictions")
        print("   3. Try opening with LibreOffice Calc or Google Sheets")
        print("   4. Clear Excel temporary files and cache")
    else:
        print("⚠️  Some tests failed. Check the error messages above.")

def test_minimal_example():
    """Test with absolute minimal data"""
    print(f"\n🧪 Minimal Example Test")
    print("-" * 30)
    
    minimal_data = [{"ID": 1, "Name": "Test"}]
    
    try:
        output = ExcelExporter.create_excel_file(minimal_data)
        content = output.getvalue()
        
        with open("minimal_test.xlsx", "wb") as f:
            f.write(content)
        
        from openpyxl import load_workbook
        wb = load_workbook("minimal_test.xlsx")
        print("✅ Minimal example works perfectly")
        print(f"   File size: {len(content)} bytes")
        print(f"   Headers: {[cell.value for cell in wb.active[1]]}")
        wb.close()
        
    except Exception as e:
        print(f"❌ Even minimal example fails: {str(e)}")

if __name__ == "__main__":
    test_excel_compatibility()
    test_minimal_example()
