"""
Test the API endpoint to download Excel files and check format
"""
import requests
import tempfile
import os
import json
from openpyxl import load_workbook

def test_api_excel_download():
    """Test the API Excel generation and download"""
    
    base_url = "http://127.0.0.1:5000"
    
    # Sample data
    test_data = []
    for i in range(10):
        test_data.append({
            "id": i + 1,
            "full_name": f"Test User {i + 1}",
            "email": f"user{i + 1}@example.com",
            "nationality": "Test Country",
            "institution": "Test University",
            "phone_number": f"+1234567890{i}",
            "category": "test",
            "form_status": "approved", 
            "payment_status": "paid",
            "created_at": "2025-08-14T10:00:00"
        })
    
    print("Testing API Excel endpoint...\n")
    
    try:
        # Test 1: Create export
        print("Step 1: Creating export via API")
        export_request = {
            "data": test_data,
            "template": "standard",
            "filename": "api_test_participants.xlsx"
        }
        
        response = requests.post(f"{base_url}/api/ybb/export/participants", 
                               json=export_request, 
                               timeout=30)
        
        if response.status_code != 200:
            print(f"❌ Export creation failed: {response.status_code}")
            print(f"Response: {response.text}")
            return
        
        result = response.json()
        if result["status"] != "success":
            print(f"❌ Export failed: {result}")
            return
        
        export_id = result["data"]["export_id"]
        filename = result["data"]["file_name"]
        print(f"✅ Export created: {export_id}")
        print(f"Filename: {filename}")
        print(f"Strategy: {result.get('export_strategy', 'unknown')}")
        
        # Step 2: Download the file
        print(f"\nStep 2: Downloading file via API")
        download_response = requests.get(f"{base_url}/api/ybb/export/{export_id}/download", 
                                      timeout=30)
        
        if download_response.status_code != 200:
            print(f"❌ Download failed: {download_response.status_code}")
            print(f"Response: {download_response.text}")
            return
        
        # Check headers
        content_type = download_response.headers.get('content-type', 'unknown')
        content_disposition = download_response.headers.get('content-disposition', 'unknown')
        content_length = download_response.headers.get('content-length', 'unknown')
        
        print(f"✅ Download successful")
        print(f"Content-Type: {content_type}")
        print(f"Content-Disposition: {content_disposition}")
        print(f"Content-Length: {content_length} bytes")
        
        # Step 3: Save and validate the file
        print(f"\nStep 3: Validating downloaded file")
        file_content = download_response.content
        
        print(f"File size: {len(file_content)} bytes")
        print(f"Header check: {file_content[:4]} (should be b'PK\\x03\\x04')")
        
        if not file_content.startswith(b'PK\x03\x04'):
            print("❌ File does not have valid Excel header!")
            print(f"First 50 bytes: {file_content[:50]}")
            return
        
        # Save to temp file with exact filename from API
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        temp_file.write(file_content)
        temp_file.close()
        
        print(f"Saved to temp file: {temp_file.name}")
        
        # Try to open with Excel library
        try:
            wb = load_workbook(temp_file.name)
            print(f"✅ Excel validation passed!")
            print(f"Sheets: {wb.sheetnames}")
            
            # Check data
            ws = wb.active
            print(f"Data dimensions: {ws.max_row} rows × {ws.max_column} columns")
            
            # Check some cell values
            if ws.max_row > 1:
                print(f"Headers (Row 1): {[ws.cell(1, col).value for col in range(1, min(6, ws.max_column + 1))]}")
                print(f"First data row: {[ws.cell(2, col).value for col in range(1, min(6, ws.max_column + 1))]}")
            
            wb.close()
            
            # Test if Windows can recognize it (try to get file info)
            file_stats = os.stat(temp_file.name)
            print(f"File stats: {file_stats.st_size} bytes, modified: {file_stats.st_mtime}")
            
            print("\n✅ ALL TESTS PASSED! File is valid Excel format.")
            
        except Exception as excel_error:
            print(f"❌ Excel validation failed: {excel_error}")
            
            # Additional debugging
            print(f"\nDebugging info:")
            print(f"File exists: {os.path.exists(temp_file.name)}")
            print(f"File size on disk: {os.path.getsize(temp_file.name)}")
            
            # Try to read as binary and check format
            with open(temp_file.name, 'rb') as f:
                file_header = f.read(100)
            print(f"File header (100 bytes): {file_header}")
        
        # Cleanup
        os.unlink(temp_file.name)
        
    except Exception as e:
        print(f"❌ Test failed with exception: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    test_api_excel_download()
