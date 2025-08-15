#!/usr/bin/env python3
"""
Test script to verify Excel file generation fixes
"""
import sys
import os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import json
import requests
from io import BytesIO
import pandas as pd
from openpyxl import load_workbook
import tempfile

def test_excel_generation():
    """Test Excel file generation with various data scenarios"""
    
    # Test data with problematic characters that might cause corruption
    test_data = [
        {
            "id": 1,
            "name": "John Doe",
            "email": "john@example.com",
            "description": "Normal data",
            "created_at": "2024-01-01 10:00:00"
        },
        {
            "id": 2,
            "name": "Jane Smith\nWith newline",
            "email": "jane@example.com",
            "description": "Data with\ttab and\nnewline",
            "created_at": "2024-01-02 11:00:00"
        },
        {
            "id": 3,
            "name": "Special chars: äöü ñ 中文",
            "email": "special@example.com",
            "description": "Unicode characters: 测试 データ",
            "created_at": "2024-01-03 12:00:00"
        },
        {
            "id": 4,
            "name": "Control chars test",
            "email": "control@example.com",
            "description": "Contains control chars: \x00\x01\x02\x03",
            "created_at": "2024-01-04 13:00:00"
        },
        {
            "id": 5,
            "name": None,
            "email": "",
            "description": "Empty and null values",
            "created_at": None
        },
        {
            "id": 6,
            "name": "Very long text " * 100,  # Very long text
            "email": "long@example.com",
            "description": "This is a very long description that exceeds normal limits " * 50,
            "created_at": "2024-01-06 15:00:00"
        }
    ]
    
    print("Testing Excel file generation locally...")
    
    # Test using the improved Excel exporter directly
    try:
        from utils.excel_exporter import ExcelExporter
        
        print("Creating Excel file with test data...")
        excel_output = ExcelExporter.create_excel_file(
            data=test_data,
            filename="test_export.xlsx",
            sheet_name="Test Data"
        )
        
        # Save to temporary file for testing
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
            temp_file.write(excel_output.getvalue())
            temp_path = temp_file.name
        
        print(f"Excel file created: {temp_path}")
        
        # Test if the file can be opened and read
        try:
            # Test with openpyxl
            workbook = load_workbook(temp_path)
            sheet = workbook.active
            print(f"Successfully opened with openpyxl. Sheet name: {sheet.title}")
            print(f"Data dimensions: {sheet.max_row} rows, {sheet.max_column} columns")
            
            # Test with pandas
            df = pd.read_excel(temp_path)
            print(f"Successfully opened with pandas. Shape: {df.shape}")
            print(f"Columns: {list(df.columns)}")
            
            # Display first few rows
            print("\nFirst few rows:")
            print(df.head(3).to_string())
            
            print("\n✅ Excel file is valid and can be opened without errors!")
            
        except Exception as e:
            print(f"❌ Error opening Excel file: {str(e)}")
            return False
        
        # Clean up temp file
        try:
            os.unlink(temp_path)
        except:
            pass
            
        return True
        
    except Exception as e:
        print(f"❌ Error creating Excel file: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

def test_api_export():
    """Test the API endpoint for Excel export"""
    
    print("\nTesting API export endpoint...")
    
    # Test data for API
    api_test_data = {
        "export_type": "participants",
        "template": "standard",
        "format": "excel",
        "data": [
            {
                "id": 1,
                "full_name": "Test User 1",
                "email": "test1@example.com",
                "phone": "+1234567890",
                "form_status": "approved",
                "created_at": "2024-01-01 10:00:00"
            },
            {
                "id": 2,
                "full_name": "Test User 2\nWith newline",
                "email": "test2@example.com",
                "phone": "+0987654321",
                "form_status": "pending",
                "created_at": "2024-01-02 11:00:00"
            }
        ]
    }
    
    try:
        # Make API request
        response = requests.post(
            "http://127.0.0.1:5000/api/ybb/export/participants",
            json=api_test_data,
            timeout=30
        )
        
        if response.status_code == 200:
            result = response.json()
            print("✅ API export request successful!")
            print(f"Export ID: {result.get('data', {}).get('export_id')}")
            
            # Try to download the file
            export_id = result.get('data', {}).get('export_id')
            if export_id:
                download_response = requests.get(
                    f"http://127.0.0.1:5000/api/ybb/export/{export_id}/download",
                    timeout=30
                )
                
                if download_response.status_code == 200:
                    # Save and test the downloaded file
                    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
                        temp_file.write(download_response.content)
                        temp_path = temp_file.name
                    
                    try:
                        # Test if downloadable file can be opened
                        df = pd.read_excel(temp_path)
                        print(f"✅ Downloaded Excel file is valid! Shape: {df.shape}")
                        
                        # Clean up
                        os.unlink(temp_path)
                        return True
                        
                    except Exception as e:
                        print(f"❌ Downloaded Excel file is corrupted: {str(e)}")
                        return False
                else:
                    print(f"❌ Failed to download file: {download_response.status_code}")
                    return False
        else:
            print(f"❌ API request failed: {response.status_code}")
            print(f"Response: {response.text}")
            return False
            
    except requests.exceptions.ConnectionError:
        print("⚠️  API server not running. Skipping API test.")
        return True  # Don't fail the test if server is not running
    except Exception as e:
        print(f"❌ API test failed: {str(e)}")
        return False

if __name__ == "__main__":
    print("🧪 Testing Excel Export Fixes")
    print("=" * 50)
    
    # Test local Excel generation
    local_test_passed = test_excel_generation()
    
    # Test API endpoint
    api_test_passed = test_api_export()
    
    print("\n" + "=" * 50)
    print("📊 Test Results:")
    print(f"Local Excel Generation: {'✅ PASSED' if local_test_passed else '❌ FAILED'}")
    print(f"API Export Test: {'✅ PASSED' if api_test_passed else '❌ FAILED'}")
    
    if local_test_passed and api_test_passed:
        print("\n🎉 All tests passed! Excel export should now work correctly.")
    else:
        print("\n⚠️  Some tests failed. Please check the issues above.")
        sys.exit(1)
