"""
Fix Excel format issues - comprehensive validation and fixing
"""
import pandas as pd
import tempfile
import os
from openpyxl import Workbook, load_workbook
from io import BytesIO
import mimetypes

def fix_excel_generation():
    """Fix Excel generation with comprehensive validation"""
    
    print("🔧 Excel Format Issue Analysis & Fix\n")
    
    # Test data
    test_data = [
        {"id": 1, "name": "John", "email": "john@test.com", "status": "active"},
        {"id": 2, "name": "Jane", "email": "jane@test.com", "status": "pending"},
        {"id": 3, "name": "Bob", "email": "bob@test.com", "status": "active"}
    ]
    
    # Issue 1: File extension validation
    print("1️⃣ Testing file extension detection...")
    test_filenames = [
        "test.xlsx",
        "test.xls", 
        "test",
        "test.XLSX",
        "test.csv.xlsx",
        "participants_export.xlsx"
    ]
    
    for filename in test_filenames:
        detected_type = mimetypes.guess_type(filename)[0] or "unknown"
        is_excel = filename.lower().endswith(('.xlsx', '.xls'))
        print(f"   {filename:<25} → MIME: {detected_type:<50} Excel: {is_excel}")
    
    print()
    
    # Issue 2: Excel creation validation with multiple methods
    print("2️⃣ Testing Excel creation methods...")
    
    methods = [
        ("pandas_basic", create_excel_pandas_basic),
        ("pandas_with_options", create_excel_pandas_with_options),
        ("openpyxl_manual", create_excel_openpyxl_manual),
        ("openpyxl_with_formatting", create_excel_openpyxl_formatted)
    ]
    
    results = {}
    
    for method_name, method_func in methods:
        try:
            print(f"   Testing {method_name}...")
            excel_data, validation_result = method_func(test_data)
            results[method_name] = {
                'success': True,
                'size': len(excel_data),
                'validation': validation_result,
                'data': excel_data
            }
            print(f"   ✅ {method_name}: {len(excel_data)} bytes, validation: {validation_result}")
            
        except Exception as e:
            results[method_name] = {
                'success': False,
                'error': str(e)
            }
            print(f"   ❌ {method_name}: {e}")
    
    print()
    
    # Issue 3: File serving validation
    print("3️⃣ Testing file serving headers...")
    
    # Test different content types
    content_types = [
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        'application/vnd.ms-excel',
        'application/octet-stream'
    ]
    
    for ct in content_types:
        print(f"   Content-Type: {ct}")
        print(f"   → Recommended for: {'XLSX' if 'openxml' in ct else 'Legacy XLS' if 'ms-excel' in ct else 'Generic'}")
    
    print()
    
    # Issue 4: Create fixed Excel exporter
    print("4️⃣ Creating robust Excel exporter...")
    
    best_method = None
    for method_name, result in results.items():
        if result['success'] and result['validation']['valid']:
            best_method = method_name
            break
    
    if best_method:
        print(f"   ✅ Best method identified: {best_method}")
        
        # Create enhanced exporter
        fixed_excel_data = results[best_method]['data']
        
        # Validate thoroughly
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        temp_file.write(fixed_excel_data)
        temp_file.close()
        
        # Multi-layer validation
        validations = validate_excel_file(temp_file.name)
        print(f"   📋 Validation Results:")
        for check, result in validations.items():
            status = "✅" if result['passed'] else "❌"
            print(f"      {status} {check}: {result['message']}")
        
        # Test Windows file association
        file_ext = os.path.splitext(temp_file.name)[1]
        detected_mime = mimetypes.guess_type(temp_file.name)[0]
        print(f"   📁 File extension: {file_ext}")
        print(f"   📄 Detected MIME: {detected_mime}")
        
        os.unlink(temp_file.name)
        
        return fixed_excel_data, validations
    else:
        print("   ❌ No working method found!")
        return None, None

def create_excel_pandas_basic(data):
    """Basic pandas Excel creation"""
    df = pd.DataFrame(data)
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Data')
    output.seek(0)
    content = output.getvalue()
    validation = validate_excel_content(content)
    return content, validation

def create_excel_pandas_with_options(data):
    """Pandas Excel with options (problematic)"""
    df = pd.DataFrame(data)
    output = BytesIO()
    # This might fail on newer pandas versions
    with pd.ExcelWriter(output, engine='openpyxl', options={'remove_timezone': True}) as writer:
        df.to_excel(writer, index=False, sheet_name='Data')
    output.seek(0)
    content = output.getvalue()
    validation = validate_excel_content(content)
    return content, validation

def create_excel_openpyxl_manual(data):
    """Manual openpyxl creation"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    
    # Headers
    headers = list(data[0].keys()) if data else []
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Data
    for row, record in enumerate(data, 2):
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=record.get(header, ''))
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    content = output.getvalue()
    validation = validate_excel_content(content)
    return content, validation

def create_excel_openpyxl_formatted(data):
    """openpyxl with formatting"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    
    # Headers with formatting
    headers = list(data[0].keys()) if data else []
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = cell.font.copy(bold=True)
    
    # Data
    for row, record in enumerate(data, 2):
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=record.get(header, ''))
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    content = output.getvalue()
    validation = validate_excel_content(content)
    return content, validation

def validate_excel_content(content):
    """Validate Excel content"""
    validation = {
        'valid': False,
        'size_ok': len(content) > 100,
        'header_ok': content.startswith(b'PK\x03\x04'),
        'excel_readable': False,
        'sheets_count': 0,
        'error': None
    }
    
    if validation['size_ok'] and validation['header_ok']:
        try:
            # Try to read with openpyxl
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
            temp_file.write(content)
            temp_file.close()
            
            wb = load_workbook(temp_file.name)
            validation['excel_readable'] = True
            validation['sheets_count'] = len(wb.sheetnames)
            validation['valid'] = True
            wb.close()
            
            os.unlink(temp_file.name)
            
        except Exception as e:
            validation['error'] = str(e)
    
    return validation

def validate_excel_file(file_path):
    """Comprehensive file validation"""
    validations = {}
    
    # File existence
    validations['file_exists'] = {
        'passed': os.path.exists(file_path),
        'message': f"File exists at {file_path}"
    }
    
    if not os.path.exists(file_path):
        return validations
    
    # File size
    file_size = os.path.getsize(file_path)
    validations['file_size'] = {
        'passed': file_size > 100,
        'message': f"File size: {file_size} bytes"
    }
    
    # File header
    with open(file_path, 'rb') as f:
        header = f.read(4)
    
    validations['excel_header'] = {
        'passed': header == b'PK\x03\x04',
        'message': f"Header: {header} (should be b'PK\\x03\\x04')"
    }
    
    # Excel readability
    try:
        wb = load_workbook(file_path)
        validations['excel_readable'] = {
            'passed': True,
            'message': f"Readable with {len(wb.sheetnames)} sheets: {wb.sheetnames}"
        }
        wb.close()
    except Exception as e:
        validations['excel_readable'] = {
            'passed': False,
            'message': f"Not readable: {e}"
        }
    
    # MIME type detection
    detected_mime = mimetypes.guess_type(file_path)[0]
    validations['mime_type'] = {
        'passed': detected_mime in ['application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', None],
        'message': f"Detected MIME: {detected_mime}"
    }
    
    return validations

if __name__ == "__main__":
    fix_excel_generation()
