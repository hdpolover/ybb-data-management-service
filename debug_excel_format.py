"""
Debug Excel format issues - test different approaches to Excel generation
"""
import pandas as pd
from openpyxl import Workbook
from io import BytesIO
import os
import tempfile

def test_excel_generation():
    """Test different Excel generation approaches"""
    
    # Sample data
    test_data = [
        {"id": 1, "name": "John Doe", "email": "john@example.com", "country": "USA"},
        {"id": 2, "name": "Jane Smith", "email": "jane@example.com", "country": "UK"},
        {"id": 3, "name": "Bob Johnson", "email": "bob@example.com", "country": "Canada"}
    ]
    
    df = pd.DataFrame(test_data)
    
    print("Testing Excel generation methods...\n")
    
    # Method 1: Basic pandas with openpyxl
    try:
        print("Method 1: Basic pandas with openpyxl")
        output1 = BytesIO()
        with pd.ExcelWriter(output1, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Data', index=False)
        
        content1 = output1.getvalue()
        print(f"✅ Success - Size: {len(content1)} bytes")
        print(f"Header check: {content1[:4]} (should be b'PK\\x03\\x04')")
        
        # Write to temp file and try to validate
        temp_file1 = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        temp_file1.write(content1)
        temp_file1.close()
        
        # Try to open with openpyxl to validate
        from openpyxl import load_workbook
        try:
            wb_test = load_workbook(temp_file1.name)
            print(f"✅ File validation passed - {len(wb_test.sheetnames)} sheets")
            wb_test.close()
        except Exception as ve:
            print(f"❌ File validation failed: {ve}")
        
        os.unlink(temp_file1.name)
        print()
        
    except Exception as e:
        print(f"❌ Failed: {e}\n")
    
    # Method 2: pandas with options that might cause issues
    try:
        print("Method 2: pandas with potentially problematic options")
        output2 = BytesIO()
        with pd.ExcelWriter(output2, engine='openpyxl', options={'remove_timezone': True}) as writer:
            df.to_excel(writer, sheet_name='Data', index=False)
        
        content2 = output2.getvalue()
        print(f"✅ Success - Size: {len(content2)} bytes")
        print(f"Header check: {content2[:4]} (should be b'PK\\x03\\x04')")
        
        temp_file2 = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        temp_file2.write(content2)
        temp_file2.close()
        
        from openpyxl import load_workbook
        try:
            wb_test = load_workbook(temp_file2.name)
            print(f"✅ File validation passed - {len(wb_test.sheetnames)} sheets")
            wb_test.close()
        except Exception as ve:
            print(f"❌ File validation failed: {ve}")
        
        os.unlink(temp_file2.name)
        print()
        
    except Exception as e:
        print(f"❌ Failed: {e}\n")
    
    # Method 3: Manual openpyxl
    try:
        print("Method 3: Manual openpyxl creation")
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        
        # Write headers
        headers = list(df.columns)
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)
        
        # Write data
        for row_num, (_, row_data) in enumerate(df.iterrows(), 2):
            for col_num, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=value)
        
        output3 = BytesIO()
        wb.save(output3)
        content3 = output3.getvalue()
        
        print(f"✅ Success - Size: {len(content3)} bytes")
        print(f"Header check: {content3[:4]} (should be b'PK\\x03\\x04')")
        
        temp_file3 = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        temp_file3.write(content3)
        temp_file3.close()
        
        from openpyxl import load_workbook
        try:
            wb_test = load_workbook(temp_file3.name)
            print(f"✅ File validation passed - {len(wb_test.sheetnames)} sheets")
            wb_test.close()
        except Exception as ve:
            print(f"❌ File validation failed: {ve}")
        
        os.unlink(temp_file3.name)
        print()
        
    except Exception as e:
        print(f"❌ Failed: {e}\n")
    
    # Method 4: Test with ExcelExporter class
    try:
        print("Method 4: Using ExcelExporter class")
        from utils.excel_exporter import ExcelExporter
        
        output4 = ExcelExporter.create_excel_file(test_data, sheet_name="Data")
        content4 = output4.getvalue()
        
        print(f"✅ Success - Size: {len(content4)} bytes")
        print(f"Header check: {content4[:4]} (should be b'PK\\x03\\x04')")
        
        temp_file4 = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        temp_file4.write(content4)
        temp_file4.close()
        
        from openpyxl import load_workbook
        try:
            wb_test = load_workbook(temp_file4.name)
            print(f"✅ File validation passed - {len(wb_test.sheetnames)} sheets")
            wb_test.close()
        except Exception as ve:
            print(f"❌ File validation failed: {ve}")
        
        os.unlink(temp_file4.name)
        print()
        
    except Exception as e:
        print(f"❌ Failed: {e}\n")
    
    print("Testing completed.")

if __name__ == "__main__":
    test_excel_generation()
