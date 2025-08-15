#!/usr/bin/env python3
"""
Comprehensive test for Excel corruption edge cases
"""
import sys
import os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import tempfile
import pandas as pd
from openpyxl import load_workbook

def test_problematic_data():
    """Test with data that commonly causes Excel corruption"""
    
    # Data with various problematic scenarios
    problematic_data = [
        {
            "id": 1,
            "name": "Normal Data",
            "email": "normal@test.com",
            "notes": "Regular text"
        },
        {
            "id": 2,
            "name": "Newlines\nand\r\ncarriage returns",
            "email": "newlines@test.com",
            "notes": "Text with\nnewlines\rand\r\ncarriage returns"
        },
        {
            "id": 3,
            "name": "Unicode: 测试 データ ñáéíóú",
            "email": "unicode@test.com", 
            "notes": "Mixed: Hello 世界 🌍 Ħěļłő"
        },
        {
            "id": 4,
            "name": "Control chars: \x00\x01\x02\x03\x04\x05\x06\x07\x08\x0B\x0C\x0E\x0F",
            "email": "control@test.com",
            "notes": "More control: \x10\x11\x12\x13\x14\x15\x16\x17\x18\x19\x1A\x1B\x1C\x1D\x1E\x1F"
        },
        {
            "id": 5,
            "name": "DEL char: \x7F and high ASCII: \x80\x81\x82",
            "email": "del@test.com",
            "notes": "Extended ASCII: \x90\x91\x92\x93\x94\x95"
        },
        {
            "id": 6,
            "name": "Very long text: " + "A" * 40000,  # Exceeds Excel cell limit
            "email": "long@test.com",
            "notes": "Another long field: " + "B" * 35000
        },
        {
            "id": 7,
            "name": "Tabs\tand\tmultiple\ttabs",
            "email": "tabs@test.com",
            "notes": "Mixed\ttabs\nand\rnewlines\r\ntogether"
        },
        {
            "id": 8,
            "name": None,  # Null value
            "email": "",   # Empty string
            "notes": None  # Another null
        },
        {
            "id": 9,
            "name": "Formula injection: =SUM(1+1)",
            "email": "formula@test.com",
            "notes": "Potential formulas: =A1+B1, @SUM(1,2)"
        },
        },
        {
            "id": 10,
            "name": "Special quotes: curly quotes and German style",
            "email": "quotes@test.com",
            "notes": "Mixed punctuation: ellipsis, dash, em-dash"
        }
    ]
    
    print("Testing with problematic data scenarios...")
    
    try:
        from utils.excel_exporter import ExcelExporter
        
        # Create Excel file
        excel_output = ExcelExporter.create_excel_file(
            data=problematic_data,
            filename="problematic_test.xlsx",
            sheet_name="Problematic Data"
        )
        
        # Save to temp file
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
            temp_file.write(excel_output.getvalue())
            temp_path = temp_file.name
        
        print(f"Created file: {temp_path}")
        
        # Test opening with different methods
        success_count = 0
        
        # Test 1: openpyxl
        try:
            workbook = load_workbook(temp_path)
            sheet = workbook.active
            print(f"✅ openpyxl: {sheet.max_row} rows, {sheet.max_column} cols")
            success_count += 1
        except Exception as e:
            print(f"❌ openpyxl failed: {str(e)}")
        
        # Test 2: pandas
        try:
            df = pd.read_excel(temp_path)
            print(f"✅ pandas: shape {df.shape}")
            success_count += 1
            
            # Check for data integrity
            print("Checking data integrity...")
            for idx, row in df.iterrows():
                if idx < 3:  # Check first few rows
                    print(f"  Row {idx}: id={row.get('id')}, name_length={len(str(row.get('name', '')))}")
            
        except Exception as e:
            print(f"❌ pandas failed: {str(e)}")
        
        # Test 3: Try opening in Excel via COM (Windows only)
        try:
            try:
                import win32com.client
                excel_app = win32com.client.Dispatch("Excel.Application")
                excel_app.Visible = False
                workbook = excel_app.Workbooks.Open(temp_path)
                sheet = workbook.ActiveSheet
                print(f"✅ Excel COM: {sheet.UsedRange.Rows.Count} rows")
                workbook.Close()
                excel_app.Quit()
                success_count += 1
            except ImportError:
                print("⚠️  Excel COM test skipped (pywin32 not available)")
        except Exception as e:
            print(f"❌ Excel COM failed: {str(e)}")
        
        # Clean up
        try:
            os.unlink(temp_path)
        except:
            pass
        
        print(f"\n📊 Results: {success_count}/2+ tests passed")
        return success_count >= 2
        
    except Exception as e:
        print(f"❌ Test failed with error: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

def test_empty_and_edge_cases():
    """Test edge cases like empty data, single row, etc."""
    
    print("\nTesting edge cases...")
    
    test_cases = [
        ("Empty data", []),
        ("Single row", [{"id": 1, "name": "Only Row"}]),
        ("Single column", [{"id": i} for i in range(5)]),
        ("All null values", [{"a": None, "b": None, "c": None} for _ in range(3)]),
        ("Mixed data types", [
            {"int": 123, "float": 45.67, "bool": True, "str": "text", "none": None},
            {"int": 0, "float": 0.0, "bool": False, "str": "", "none": None}
        ])
    ]
    
    passed = 0
    total = len(test_cases)
    
    for name, data in test_cases:
        try:
            print(f"  Testing: {name}")
            
            if not data:  # Skip empty data test as it should raise ValueError
                try:
                    from utils.excel_exporter import ExcelExporter
                    ExcelExporter.create_excel_file(data=data)
                    print(f"    ❌ Expected ValueError for empty data")
                except ValueError:
                    print(f"    ✅ Correctly handled empty data")
                    passed += 1
                except Exception as e:
                    print(f"    ❌ Unexpected error: {str(e)}")
                continue
            
            from utils.excel_exporter import ExcelExporter
            excel_output = ExcelExporter.create_excel_file(data=data)
            
            # Quick validation
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
                temp_file.write(excel_output.getvalue())
                temp_path = temp_file.name
            
            df = pd.read_excel(temp_path)
            print(f"    ✅ Success: {df.shape}")
            passed += 1
            
            os.unlink(temp_path)
            
        except Exception as e:
            print(f"    ❌ Failed: {str(e)}")
    
    print(f"\nEdge cases: {passed}/{total} passed")
    return passed == total

if __name__ == "__main__":
    print("🔍 Comprehensive Excel Corruption Tests")
    print("=" * 50)
    
    test1 = test_problematic_data()
    test2 = test_empty_and_edge_cases()
    
    print("\n" + "=" * 50)
    print("📊 Final Results:")
    print(f"Problematic Data Test: {'✅ PASSED' if test1 else '❌ FAILED'}")
    print(f"Edge Cases Test: {'✅ PASSED' if test2 else '❌ FAILED'}")
    
    if test1 and test2:
        print("\n🎉 All comprehensive tests passed!")
        print("✨ Excel files should now be corruption-free!")
    else:
        print("\n⚠️  Some tests failed. Check the output above.")
        sys.exit(1)
