#!/usr/bin/env python3
"""
Test script to replicate the exact Excel corruption issue
"""
import sys
import os
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from services.ybb_export_service import YBBExportService
import logging

# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def test_exact_filename():
    """Test with the exact filename from the error message"""
    
    # Create test data that might cause issues
    test_data = [
        {
            'id': 1,
            'full_name': 'Test User 1',
            'email': 'test1@example.com',
            'status': 'Approved',
            'registration_date': '2025-07-27',
            'country': 'Japan',
            'institution': 'Tokyo University',
            'phone_number': '+81-90-1234-5678',
            'category': 'Student',
            'notes': 'Regular participant',
            'payment_status': 'Completed'
        },
        {
            'id': 2, 
            'full_name': 'Test User 2',
            'email': 'test2@example.com',
            'status': 'Approved',
            'registration_date': '2025-07-26',
            'country': 'Korea',
            'institution': 'Seoul National University', 
            'phone_number': '+82-10-9876-5432',
            'category': 'Professional',
            'notes': 'VIP participant',
            'payment_status': 'Pending'
        }
    ]
    
    # Create export request that would generate the problematic filename
    export_request = {
        'export_type': 'participants',
        'template': 'standard', 
        'data': test_data,
        'format': 'excel',
        'event_name': 'Japan Youth Summit 2025 Participants',
        'status_filter': 'Approved Forms',
        'export_date': '27-07-2025'
    }
    
    export_service = YBBExportService()
    
    try:
        print("Creating export with potentially problematic filename...")
        result = export_service.create_export(export_request)
        
        if result['status'] == 'success':
            export_id = result['data']['export_id']
            print(f"Export created successfully: {export_id}")
            
            # Download the file
            file_content, filename = export_service.download_export(export_id, 'single')
            
            if file_content:
                # Use the exact filename from the error message
                problem_filename = "Japan_Youth_Summit_2025_Participants_Participants_Approved_Forms_27-07-2025 (2).xlsx"
                
                print(f"Generated filename: {filename}")
                print(f"Problem filename: {problem_filename}")
                print(f"File size: {len(file_content)} bytes")
                
                # Test both filenames
                for test_filename in [filename, problem_filename]:
                    try:
                        print(f"\nTesting filename: {test_filename}")
                        
                        # Write the file
                        with open(test_filename, 'wb') as f:
                            f.write(file_content)
                        
                        print(f"File written successfully: {os.path.getsize(test_filename)} bytes")
                        
                        # Validate with openpyxl
                        from openpyxl import load_workbook
                        
                        wb = load_workbook(test_filename)
                        ws = wb.active
                        
                        print(f"✅ File validation PASSED")
                        print(f"   Sheet: {ws.title}")
                        print(f"   Dimensions: {ws.max_row} rows x {ws.max_column} columns")
                        
                        if ws.max_row > 1:
                            print(f"   Sample data: {[cell.value for cell in ws[2][:3]]}")
                        
                        wb.close()
                        
                        # Check file properties
                        stat = os.stat(test_filename)
                        print(f"   File size on disk: {stat.st_size} bytes")
                        print(f"   Modified: {stat.st_mtime}")
                        
                    except Exception as file_error:
                        print(f"❌ File test FAILED for {test_filename}: {str(file_error)}")
                        
                        # Analyze the file content
                        print(f"   File content analysis:")
                        print(f"   - First 20 bytes: {file_content[:20].hex()}")
                        print(f"   - Last 20 bytes: {file_content[-20:].hex()}")
                        print(f"   - Content length: {len(file_content)}")
                        print(f"   - Starts with PK: {file_content.startswith(b'PK')}")
                        
                # Test with a simple filename to compare
                simple_filename = "test_simple.xlsx"
                with open(simple_filename, 'wb') as f:
                    f.write(file_content)
                
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(simple_filename)
                    print(f"\n✅ Simple filename test PASSED: {simple_filename}")
                    wb.close()
                except Exception as e:
                    print(f"\n❌ Simple filename test FAILED: {str(e)}")
                    
            else:
                print("❌ No file content returned from export service")
                
        else:
            print(f"❌ Export creation failed: {result.get('message', 'Unknown error')}")
            
    except Exception as e:
        print(f"❌ Test failed with exception: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    test_exact_filename()
