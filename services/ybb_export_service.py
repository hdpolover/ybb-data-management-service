"""
YBB Export Service - Handles YBB-specific data export logic
"""
import uuid
from datetime import datetime, timedelta
import os
import tempfile
import zipfile
from io import BytesIO
import logging
import time
from openpyxl import Workbook

# Try to import pandas gracefully
try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError as e:
    print(f"⚠️  pandas not available in ybb_export_service: {e}")
    PANDAS_AVAILABLE = False
    
    # Create a comprehensive dummy pandas for graceful degradation
    class DummyDataFrame:
        def __init__(self, *args, **kwargs):
            raise ImportError("pandas is not available - C++ libraries missing")
        
        def __getattr__(self, name):
            raise ImportError(f"pandas DataFrame not available - cannot access '{name}' method")
    
    class DummySeries:
        def __init__(self, *args, **kwargs):
            raise ImportError("pandas Series not available - C++ libraries missing")
        
        def __getattr__(self, name):
            raise ImportError(f"pandas Series not available - cannot access '{name}' method")
    
    class DummyPandas:
        def __init__(self):
            self.DataFrame = DummyDataFrame
            self.Series = DummySeries
        
        def __getattr__(self, name):
            # Handle common pandas functions that might be called
            if name in ['isna', 'isnull', 'notna', 'notnull', 'read_excel', 'read_csv', 
                       'concat', 'merge', 'to_datetime', 'cut', 'qcut', 'pivot_table']:
                def pandas_not_available(*args, **kwargs):
                    raise ImportError(f"pandas is not available - cannot use pandas.{name}()")
                return pandas_not_available
            raise ImportError(f"pandas is not available - cannot access pandas.{name}")
        
        @property 
        def __version__(self):
            return "pandas-not-available"
    
    pd = DummyPandas()

# Optional memory tracking
try:
    import psutil
    PSUTIL_AVAILABLE = True
except ImportError:
    PSUTIL_AVAILABLE = False
from openpyxl.styles import Font, PatternFill, Alignment
from config.ybb_export_config import (
    EXPORT_TEMPLATES, STATUS_MAPPINGS, SYSTEM_CONFIG, 
    get_template, get_status_label, get_chunk_size, should_use_chunked_processing,
    get_cleanup_config, get_storage_limits
)
from utils.file_manager import ExportFileManager

logger = logging.getLogger(__name__)

class YBBExportService:
    """Main service for handling YBB data exports"""
    
    def __init__(self):
        self.temp_dir = tempfile.gettempdir()
        self.exports_storage = {}  # In production, use Redis or database
        self.file_manager = ExportFileManager()
        
        # Get cleanup configuration from config file
        cleanup_config = get_cleanup_config()
        storage_limits = get_storage_limits()
        
        self.max_concurrent_exports = storage_limits.get('max_concurrent_exports', 20)
        self.auto_cleanup_enabled = cleanup_config.get('auto_cleanup_enabled', True)
        self.cleanup_on_startup = cleanup_config.get('cleanup_on_startup', True)
        self.cleanup_on_export = cleanup_config.get('cleanup_on_export', False)
        self.cleanup_interval_minutes = cleanup_config.get('cleanup_interval_minutes', 30)
        self.min_export_age_minutes = cleanup_config.get('min_export_age_minutes', 10)
        
        # Track last cleanup time
        self.last_cleanup_time = datetime.now()
        
        # Perform startup cleanup if enabled
        if self.cleanup_on_startup:
            self._cleanup_old_exports()
    
    def create_export(self, export_request):
        """Main export creation method"""
        try:
            # Check if pandas is available for data processing
            if not PANDAS_AVAILABLE:
                return {
                    "status": "error", 
                    "message": "pandas is not available - Excel export functionality requires pandas with C++ libraries. Service is running in limited mode.",
                    "error_code": "PANDAS_UNAVAILABLE"
                }
            
            # Periodic cleanup based on time interval (not before every export)
            if self.auto_cleanup_enabled and self._should_run_cleanup():
                self._cleanup_old_exports()
            
            # Validate request
            validation_result = self._validate_export_request(export_request)
            if not validation_result["valid"]:
                return {"status": "error", "message": validation_result["message"]}
            
            export_id = str(uuid.uuid4())
            export_type = export_request["export_type"]
            template_name = export_request.get("template", "standard")
            
            # Get template configuration
            template_config = get_template(export_type, template_name)
            if not template_config:
                return {"status": "error", "message": f"Template '{template_name}' not found for {export_type}"}
            
            # Process data based on size
            data = export_request["data"]
            record_count = len(data)
            
            # Check if chunking should be forced
            force_chunking = export_request.get("force_chunking", False)
            force_chunk_size = export_request.get("chunk_size")
            
            if force_chunking or should_use_chunked_processing(record_count, template_config):
                # Override chunk size if specified
                if force_chunk_size and force_chunk_size > 0:
                    template_config = template_config.copy()
                    template_config["recommended_chunk_size"] = force_chunk_size
                return self._create_large_export(export_id, export_request, template_config)
            else:
                return self._create_standard_export(export_id, export_request, template_config)
                
        except Exception as e:
            logger.error(f"Export creation failed: {str(e)}")
            return {"status": "error", "message": f"Export failed: {str(e)}"}
    
    def _validate_export_request(self, request):
        """Validate export request structure"""
        required_fields = ["export_type", "data"]
        
        for field in required_fields:
            if field not in request:
                return {"valid": False, "message": f"Missing required field: {field}"}
        
        valid_export_types = ["participants", "payments", "ambassadors"]
        if request["export_type"] not in valid_export_types:
            return {"valid": False, "message": f"Invalid export_type. Must be one of: {valid_export_types}"}
        
        if not isinstance(request["data"], list):
            return {"valid": False, "message": "Data must be an array/list"}
        
        # Validate filename parameters using file manager
        filename_validation = self.file_manager.validate_filename_params(request)
        if not filename_validation['valid']:
            return filename_validation
        
        return {"valid": True}
    
    def _get_memory_usage(self):
        """Get current memory usage in MB"""
        if PSUTIL_AVAILABLE:
            try:
                process = psutil.Process()
                memory_info = process.memory_info()
                return round(memory_info.rss / (1024 * 1024), 2)  # RSS in MB
            except Exception:
                return None
        return None
    
    def _create_standard_export(self, export_id, export_request, template_config):
        """Create standard single-file export"""
        try:
            start_time = time.time()
            start_memory = self._get_memory_usage()
            
            data = export_request["data"]
            export_type = export_request["export_type"]
            template_name = export_request.get("template", "standard")
            format_type = export_request.get("format", "excel")
            
            # Transform data
            processed_data = self._transform_data(data, export_type, template_config)
            
            # Create file
            if format_type.lower() == "excel":
                file_content, filename = self._create_excel_file(
                    processed_data, export_type, template_name, export_request
                )
            else:
                file_content, filename = self._create_csv_file(
                    processed_data, export_type, template_name, export_request
                )
            
            # Calculate comprehensive processing metrics
            total_processing_time = time.time() - start_time
            processing_time_ms = round(total_processing_time * 1000, 2)
            end_memory = self._get_memory_usage()
            memory_used = round(end_memory - start_memory, 2) if start_memory and end_memory else None
            peak_memory = end_memory if end_memory else None
            
            file_size = len(file_content)
            file_size_mb = round(file_size / (1024 * 1024), 2)
            records_per_second = round(len(data) / total_processing_time, 1) if total_processing_time > 0 else 0
            
            # Store export info
            export_info = {
                "export_id": export_id,
                "status": "success",
                "export_type": export_type,
                "template": template_name,
                "format": format_type,
                "record_count": len(data),
                "file_content": file_content,
                "filename": filename,
                "file_size": file_size,
                "file_size_mb": file_size_mb,
                "processing_time_ms": processing_time_ms,
                "memory_used_mb": memory_used,
                "peak_memory_mb": peak_memory,
                "created_at": datetime.now(),
                "expires_at": datetime.now() + timedelta(days=SYSTEM_CONFIG["limits"]["file_retention_days"])
            }
            
            self.exports_storage[export_id] = export_info
            
            return {
                "status": "success",
                "message": "Export completed successfully",
                "export_strategy": "single_file",
                "data": {
                    "export_id": export_id,
                    "file_name": filename,
                    "file_size": file_size,
                    "file_size_mb": file_size_mb,
                    "record_count": len(data),
                    "download_url": f"/api/ybb/export/{export_id}/download",
                    "expires_at": export_info["expires_at"].isoformat()
                },
                "performance_metrics": {
                    "total_processing_time_seconds": round(total_processing_time, 2),
                    "processing_time_ms": processing_time_ms,
                    "records_per_second": records_per_second,
                    "memory_used_mb": memory_used,
                    "peak_memory_mb": peak_memory,
                    "efficiency_metrics": {
                        "kb_per_record": round(file_size / len(data) / 1024, 2) if len(data) > 0 else 0,
                        "processing_ms_per_record": round(processing_time_ms / len(data), 2) if len(data) > 0 else 0,
                        "memory_efficiency_kb_per_record": round((memory_used * 1024) / len(data), 2) if memory_used and len(data) > 0 else None
                    }
                },
                "system_info": {
                    "export_type": export_type,
                    "template": template_name,
                    "format": format_type,
                    "filters_applied": export_request.get("filters", {}),
                    "generated_at": export_info["created_at"].isoformat(),
                    "compression_used": "none",
                    "temp_files_cleanup_scheduled": False
                }
            }
            
        except Exception as e:
            logger.error(f"Standard export creation failed: {str(e)}")
            raise
    
    def _create_large_export(self, export_id, export_request, template_config):
        """Create large export with chunking and performance tracking"""
        try:
            # Performance tracking
            start_time = time.time()
            chunk_start_times = []
            chunk_processing_times = []
            chunk_sizes_bytes = []
            memory_peaks = []
            
            data = export_request["data"]
            export_type = export_request["export_type"]
            template_name = export_request.get("template", "standard")
            record_count = len(data)
            
            # Track data preparation time
            prep_start = time.time()
            
            # Calculate chunk size
            chunk_size = get_chunk_size(export_type, template_name, record_count)
            
            # Split data into chunks
            chunks = [data[i:i + chunk_size] for i in range(0, len(data), chunk_size)]
            total_chunks = len(chunks)
            
            prep_time = time.time() - prep_start
            logger.info(f"Data preparation completed in {prep_time:.2f}s for {record_count} records")
            
            # Create individual files
            files_info = []
            temp_files = []
            total_uncompressed_size = 0
            
            logger.info(f"Starting chunk processing: {total_chunks} chunks of {chunk_size} records each")
            
            for i, chunk in enumerate(chunks):
                chunk_start = time.time()
                chunk_start_times.append(chunk_start)
                
                # Track memory before processing
                try:
                    import psutil
                    process = psutil.Process()
                    memory_before = process.memory_info().rss / 1024 / 1024  # MB
                except ImportError:
                    memory_before = None
                
                processed_chunk = self._transform_data(chunk, export_type, template_config)
                
                # Create chunk file
                file_content, chunk_filename = self._create_excel_file(
                    processed_chunk, export_type, template_name, export_request, 
                    batch_info={"number": i + 1, "total": total_chunks}
                )
                
                # Track processing time for this chunk
                chunk_time = time.time() - chunk_start
                chunk_processing_times.append(chunk_time)
                
                # Track file size
                chunk_size_bytes = len(file_content)
                chunk_sizes_bytes.append(chunk_size_bytes)
                total_uncompressed_size += chunk_size_bytes
                
                # Track memory after processing
                try:
                    if memory_before:
                        memory_after = process.memory_info().rss / 1024 / 1024  # MB
                        memory_peak = memory_after - memory_before
                        memory_peaks.append(memory_peak)
                except:
                    memory_peaks.append(0)
                
                # Save to temp file
                temp_file_path = os.path.join(self.temp_dir, f"{export_id}_chunk_{i+1}.xlsx")
                with open(temp_file_path, 'wb') as f:
                    f.write(file_content)
                
                temp_files.append(temp_file_path)
                
                files_info.append({
                    "batch_number": i + 1,
                    "file_name": chunk_filename,
                    "file_path": temp_file_path,
                    "file_size": chunk_size_bytes,
                    "record_count": len(chunk),
                    "record_range": f"{i * chunk_size + 1}-{min((i + 1) * chunk_size, record_count)}",
                    "processing_time_seconds": round(chunk_time, 2),
                    "records_per_second": round(len(chunk) / chunk_time, 1) if chunk_time > 0 else 0
                })
                
                logger.info(f"Chunk {i+1}/{total_chunks} completed in {chunk_time:.2f}s "
                           f"({len(chunk)} records, {chunk_size_bytes/1024:.1f}KB)")
            
            # Compression phase tracking
            compression_start = time.time()
            logger.info("Starting ZIP compression...")
            
            # Create ZIP archive
            zip_filename = self.file_manager.generate_zip_filename(export_request, export_id)
            zip_path = os.path.join(self.temp_dir, f"{export_id}_complete.zip")
            
            with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED, compresslevel=6) as zipf:
                for file_info in files_info:
                    zipf.write(file_info["file_path"], file_info["file_name"])
            
            # Get ZIP file size and calculate compression
            zip_size = os.path.getsize(zip_path)
            compression_time = time.time() - compression_start
            compression_ratio = ((total_uncompressed_size - zip_size) / total_uncompressed_size * 100) if total_uncompressed_size > 0 else 0
            
            logger.info(f"ZIP compression completed in {compression_time:.2f}s, "
                       f"ratio: {compression_ratio:.1f}%, final size: {zip_size/1024:.1f}KB")
            
            # Calculate comprehensive performance metrics
            total_processing_time = time.time() - start_time
            avg_chunk_time = sum(chunk_processing_times) / len(chunk_processing_times) if chunk_processing_times else 0
            total_records_per_second = record_count / total_processing_time if total_processing_time > 0 else 0
            avg_memory_peak = sum(memory_peaks) / len(memory_peaks) if memory_peaks else 0
            
            # Store export info
            export_info = {
                "export_id": export_id,
                "status": "success",
                "export_strategy": "multi_file",
                "export_type": export_type,
                "template": template_name,
                "record_count": record_count,
                "total_chunks": total_chunks,
                "files_info": files_info,
                "zip_path": zip_path,
                "zip_filename": zip_filename,
                "zip_size": zip_size,
                "total_file_size": total_uncompressed_size,
                "compression_ratio": compression_ratio,
                "processing_time_ms": round(total_processing_time * 1000, 2),
                "temp_files": temp_files,
                "created_at": datetime.now(),
                "expires_at": datetime.now() + timedelta(days=SYSTEM_CONFIG["limits"]["file_retention_days"])
            }
            
            self.exports_storage[export_id] = export_info
            
            return {
                "status": "success",
                "message": "Large export completed successfully",
                "export_strategy": "multi_file",
                "data": {
                    "export_id": export_id,
                    "total_records": record_count,
                    "total_files": total_chunks,
                    "individual_files": [
                        {
                            "batch_number": f["batch_number"],
                            "file_name": f["file_name"],
                            "file_size": f["file_size"],
                            "record_count": f["record_count"],
                            "record_range": f["record_range"],
                            "processing_time_seconds": f["processing_time_seconds"],
                            "records_per_second": f["records_per_second"]
                        }
                        for f in files_info
                    ],
                    "archive_info": {
                        "filename": zip_filename,
                        "compressed_size": zip_size,
                        "uncompressed_size": total_uncompressed_size,
                        "compression_ratio": f"{compression_ratio:.1f}%",
                        "compression_time_seconds": round(compression_time, 2)
                    },
                    "performance_metrics": {
                        "total_processing_time_seconds": round(total_processing_time, 2),
                        "data_preparation_time_seconds": round(prep_time, 2),
                        "average_chunk_processing_time_seconds": round(avg_chunk_time, 2),
                        "total_records_per_second": round(total_records_per_second, 1),
                        "chunk_processing_times": [round(t, 2) for t in chunk_processing_times],
                        "average_memory_peak_mb": round(avg_memory_peak, 1) if avg_memory_peak > 0 else None,
                        "efficiency_metrics": {
                            "kb_per_record_uncompressed": round(total_uncompressed_size / record_count / 1024, 2) if record_count > 0 else 0,
                            "kb_per_record_compressed": round(zip_size / record_count / 1024, 2) if record_count > 0 else 0,
                            "processing_ms_per_record": round((total_processing_time * 1000) / record_count, 2) if record_count > 0 else 0,
                            "compression_efficiency": f"{compression_ratio:.1f}%"
                        }
                    },
                    "system_info": {
                        "chunk_size": chunk_size,
                        "compression_level": 6,
                        "temp_files_cleanup_scheduled": True,
                        "export_expires_at": export_info["expires_at"].isoformat()
                    }
                },
                "download_url": f"/api/ybb/export/{export_id}/download"
            }
            
        except Exception as e:
            logger.error(f"Large export creation failed: {str(e)}")
            raise
    
    def _transform_data(self, data, export_type, template_config):
        """Transform raw data according to template configuration with sanitization"""
        if not data:
            return []
        
        fields = template_config.get("fields", [])
        headers = template_config.get("headers", fields)
        
        transformed_data = []
        
        for record in data:
            transformed_record = {}
            
            for i, field in enumerate(fields):
                header = headers[i] if i < len(headers) else field
                value = record.get(field, "N/A")
                
                # Apply transformations based on field type
                if field in ["created_at", "updated_at", "payment_date", "birthdate"]:
                    if value and value != "N/A":
                        try:
                            if isinstance(value, str):
                                # Try to parse different date formats
                                for fmt in ["%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y"]:
                                    try:
                                        parsed_date = datetime.strptime(value, fmt)
                                        value = parsed_date.strftime("%Y-%m-%d")
                                        break
                                    except ValueError:
                                        continue
                        except:
                            value = str(value) if value else ""
                
                elif field in ["amount", "usd_amount"]:
                    if value and value != "N/A":
                        try:
                            value = f"{float(value):,.2f}"
                        except:
                            value = str(value) if value else ""
                
                elif field in ["form_status", "status"]:
                    if export_type == "participants":
                        value = get_status_label("form_status", value)
                    elif export_type == "payments":
                        value = get_status_label("payment_status", value)
                
                elif field == "payment_method_id":
                    value = get_status_label("payment_method", value)
                
                elif field in ["is_active", "is_deleted"]:
                    value = get_status_label("boolean_status", value)
                
                elif field == "category":
                    value = get_status_label("category", value)
                
                # Ensure value is not None and convert to string
                if value is None:
                    value = ""
                else:
                    value = str(value)
                
                # Clean and sanitize the value for Excel compatibility
                value = self._sanitize_excel_value_enhanced(value)
                
                transformed_record[header] = value
            
            transformed_data.append(transformed_record)
        
        return transformed_data
    
    def _sanitize_excel_value_enhanced(self, value):
        """Enhanced Excel value sanitization"""
        # Handle None values
        if value is None:
            return ""
        
        # Handle pandas NaN values if pandas is available
        if PANDAS_AVAILABLE:
            try:
                if pd.isna(value):
                    return ""
            except (ImportError, AttributeError):
                # Fallback if pd.isna fails
                pass
        
        # Convert to string
        str_value = str(value)
        
        # Handle empty strings
        if not str_value.strip():
            return ""
        
        # Handle Excel formula injection (security)
        if str_value.strip().startswith(('=', '+', '-', '@', '\t', '\r', '\n')):
            str_value = "'" + str_value  # Prefix with apostrophe
        
        # Remove control characters except allowed ones
        cleaned_value = ""
        for char in str_value:
            char_code = ord(char)
            if char_code < 32:
                if char_code in [9, 10, 13]:  # Tab, LF, CR
                    cleaned_value += char
                else:
                    cleaned_value += " "  # Replace with space
            elif char_code == 127:  # DEL character
                cleaned_value += " "
            elif char_code > 1114111:  # Beyond Unicode range
                cleaned_value += " "
            else:
                cleaned_value += char
        
        # Excel cell limit (32,767 characters)
        if len(cleaned_value) > 32767:
            cleaned_value = cleaned_value[:32764] + "..."
        
        # Normalize whitespace
        import re
        cleaned_value = re.sub(r'\s+', ' ', cleaned_value).strip()
        
        return cleaned_value if cleaned_value else ""
    
    def _sanitize_sheet_name(self, sheet_name):
        """Sanitize Excel sheet name"""
        if not sheet_name:
            return "Data"
        
        # Remove invalid characters for Excel sheet names
        import re
        sanitized = re.sub(r'[\\/*\[\]:?]', '_', str(sheet_name))
        
        # Limit to Excel's 31 character limit
        if len(sanitized) > 31:
            sanitized = sanitized[:31]
        
        # Ensure not empty after sanitization
        return sanitized if sanitized.strip() else "Data"
        """
        Sanitize a single value for Excel compatibility
        
        Args:
            value: Raw value to sanitize
            
        Returns:
            Sanitized value safe for Excel
        """
        if value is None:
            return ""
        
        # Convert to string if not already
        str_value = str(value)
        
        # Handle empty strings
        if not str_value.strip():
            return ""
        
        # Remove problematic characters for Excel
        # Excel doesn't support characters below ASCII 32 except tab (9), newline (10), carriage return (13)
        cleaned_value = ""
        for char in str_value:
            char_code = ord(char)
            if char_code < 32:
                if char_code in [9, 10, 13]:  # Tab, newline, carriage return
                    cleaned_value += char
                else:
                    # Replace with space for other control characters
                    cleaned_value += " "
            elif char_code == 127:  # DEL character
                cleaned_value += " "
            else:
                cleaned_value += char
        
        # Remove excessive whitespace
        import re
        cleaned_value = re.sub(r'\s+', ' ', cleaned_value).strip()
        
        # Excel cell limit (32,767 characters)
        if len(cleaned_value) > 32767:
            cleaned_value = cleaned_value[:32764] + "..."
        
        return cleaned_value
    
    def _create_excel_file(self, data, export_type, template_name, export_request, batch_info=None):
        """Create Excel file from processed data with robust error handling and validation"""
        if not data:
            raise ValueError("No data to export")
        
        try:
            # Import the robust Excel service
            from robust_excel_service import RobustExcelService
            
            # Generate filename using file manager
            base_filename = self.file_manager.generate_filename(export_request, str(uuid.uuid4())[:8], batch_info)
            base_filename = self.file_manager.sanitize_filename(base_filename)
            
            # Get sheet name using file manager
            sheet_name = self.file_manager.get_sheet_name(export_request, batch_info)
            
            # Create Excel file using robust service with validation
            file_content, validated_filename, validation_info = RobustExcelService.create_excel_file_robust(
                data=data,
                filename=base_filename,
                sheet_name=sheet_name
            )
            
            if not validation_info['valid']:
                raise Exception(f"Excel validation failed: {validation_info['errors']}")
            
            logger.info(f"Excel file created successfully: {validated_filename} ({len(file_content)} bytes)")
            logger.debug(f"Excel validation: {validation_info}")
            
            return file_content, validated_filename
            
        except Exception as e:
            logger.error(f"Robust Excel file creation failed: {str(e)}")
            
            # Enhanced fallback with comprehensive error handling
            try:
                logger.info("Attempting enhanced fallback Excel creation")
                
                # Create DataFrame and thoroughly sanitize data
                import pandas as pd
                df = pd.DataFrame(data)
                
                # Comprehensive data cleaning
                for col in df.columns:
                    # Convert to string and sanitize
                    df[col] = df[col].astype(str).apply(lambda x: self._sanitize_excel_value_enhanced(x))
                
                # Generate and fix filename
                filename = self.file_manager.generate_filename(export_request, str(uuid.uuid4())[:8], batch_info)
                filename = self.file_manager.sanitize_filename(filename)
                
                # Ensure .xlsx extension
                if not filename.lower().endswith('.xlsx'):
                    filename = f"{os.path.splitext(filename)[0]}.xlsx"
                
                # Get sheet name and fix it
                sheet_name = self.file_manager.get_sheet_name(export_request, batch_info)
                sheet_name = self._sanitize_sheet_name(sheet_name)
                
                # Create Excel file with enhanced options
                output = BytesIO()
                try:
                    # Try basic pandas approach first
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    output.seek(0)
                    content = output.getvalue()
                    
                    # Validate the fallback file
                    if len(content) < 100 or not content.startswith(b'PK\x03\x04'):
                        raise Exception("Fallback created invalid Excel file")
                    
                    logger.info(f"Enhanced fallback Excel creation successful: {filename} ({len(content)} bytes)")
                    return content, filename
                    
                except Exception as pandas_error:
                    logger.error(f"Pandas fallback failed: {pandas_error}")
                    
                    # Ultimate fallback: Manual openpyxl
                    logger.info("Attempting manual openpyxl fallback")
                    
                    from openpyxl import Workbook
                    wb = Workbook()
                    ws = wb.active
                    ws.title = sheet_name[:31]  # Excel limit
                    
                    # Write headers
                    headers = list(df.columns)
                    for col_num, header in enumerate(headers, 1):
                        ws.cell(row=1, column=col_num, value=str(header)[:255])  # Excel cell limit
                    
                    # Write data with strict limits
                    for row_num, (_, row_data) in enumerate(df.iterrows(), 2):
                        for col_num, value in enumerate(row_data, 1):
                            cell_value = str(value)[:32767] if value else ""  # Excel cell limit
                            ws.cell(row=row_num, column=col_num, value=cell_value)
                    
                    # Save to bytes
                    manual_output = BytesIO()
                    wb.save(manual_output)
                    manual_output.seek(0)
                    manual_content = manual_output.getvalue()
                    
                    # Final validation
                    if len(manual_content) < 100 or not manual_content.startswith(b'PK\x03\x04'):
                        raise Exception("Manual openpyxl fallback created invalid file")
                    
                    logger.info(f"Manual openpyxl fallback successful: {filename} ({len(manual_content)} bytes)")
                    return manual_content, filename
                
            except Exception as fallback_error:
                logger.error(f"All Excel creation fallbacks failed: {str(fallback_error)}")
                raise Exception(f"Excel file creation completely failed. Original error: {str(e)}. Fallback error: {str(fallback_error)}")
    
    def _create_csv_file(self, data, export_type, template_name, export_request):
        """Create CSV file from processed data with enhanced error handling"""
        try:
            # Create DataFrame and sanitize data
            df = pd.DataFrame(data)
            
            # Basic data cleaning for CSV
            for col in df.columns:
                df[col] = df[col].astype(str).apply(lambda x: self._sanitize_csv_value(x))
            
            # Generate filename using file manager
            filename = self.file_manager.generate_filename(export_request, str(uuid.uuid4())[:8])
            filename = self.file_manager.sanitize_filename(filename.replace('.xlsx', '.csv'))
            
            output = BytesIO()
            df.to_csv(output, index=False, encoding='utf-8')
            output.seek(0)
            
            return output.getvalue(), filename
            
        except Exception as e:
            logger.error(f"CSV file creation failed: {str(e)}")
            raise Exception(f"CSV file creation failed: {str(e)}")
    
    def _sanitize_csv_value(self, value):
        """
        Sanitize a single value for CSV compatibility
        
        Args:
            value: Raw value to sanitize
            
        Returns:
            Sanitized value safe for CSV
        """
        if value is None:
            return ""
        
        # Convert to string if not already
        str_value = str(value)
        
        # Handle empty strings
        if not str_value.strip():
            return ""
        
        # Remove problematic characters for CSV
        import re
        cleaned_value = re.sub(r'[\r\n]+', ' ', str_value)  # Replace newlines with spaces
        cleaned_value = cleaned_value.strip()
        
        return cleaned_value
    
    def get_export_status(self, export_id):
        """Get export status and download information with detailed metrics"""
        if export_id not in self.exports_storage:
            return {"status": "error", "message": "Export not found"}
        
        export_info = self.exports_storage[export_id]
        
        # Check if expired
        if datetime.now() > export_info["expires_at"]:
            self._cleanup_export(export_id)
            return {"status": "error", "message": "Export has expired"}
        
        # Basic status info
        status_info = {
            "status": export_info["status"],
            "export_id": export_id,
            "export_type": export_info["export_type"],
            "template": export_info["template"],
            "record_count": export_info["record_count"],
            "created_at": export_info["created_at"].isoformat(),
            "expires_at": export_info["expires_at"].isoformat()
        }
        
        # Add file size information
        if "file_size" in export_info:
            status_info["file_size_bytes"] = export_info["file_size"]
            status_info["file_size_mb"] = export_info.get("file_size_mb", round(export_info["file_size"] / (1024 * 1024), 2))
        elif "zip_size" in export_info:
            status_info["zip_file_size_bytes"] = export_info["zip_size"]
            status_info["zip_file_size_mb"] = round(export_info["zip_size"] / (1024 * 1024), 2)
            if "total_file_size" in export_info:
                status_info["total_uncompressed_size_bytes"] = export_info["total_file_size"]
                status_info["total_uncompressed_size_mb"] = round(export_info["total_file_size"] / (1024 * 1024), 2)
                status_info["compression_ratio_percent"] = export_info.get("compression_ratio", 0)
        
        # Add processing time information
        if "processing_time_ms" in export_info:
            status_info["processing_time_ms"] = export_info["processing_time_ms"]
            status_info["processing_time_seconds"] = round(export_info["processing_time_ms"] / 1000, 3)
            if export_info["processing_time_ms"] > 0:
                status_info["records_per_second"] = round(export_info["record_count"] / (export_info["processing_time_ms"] / 1000), 2)
        
        # Add multi-file specific information
        if export_info.get("export_strategy") == "multi_file":
            status_info["export_strategy"] = "multi_file"
            status_info["total_chunks"] = export_info.get("total_chunks", 0)
            status_info["files_info"] = export_info.get("files_info", [])
        
        return status_info
    
    def download_export(self, export_id, file_type="single"):
        """Download export file(s)"""
        if export_id not in self.exports_storage:
            return None, None

        export_info = self.exports_storage[export_id]

        # For multi-file exports, automatically use ZIP if single file requested
        if file_type == "single" and "zip_path" in export_info and "file_content" not in export_info:
            file_type = "zip"

        if file_type == "zip" and "zip_path" in export_info:
            with open(export_info["zip_path"], 'rb') as f:
                return f.read(), export_info["zip_filename"]

        elif file_type == "single" and "file_content" in export_info:
            return export_info["file_content"], export_info["filename"]

        return None, None
    
    def download_batch_file(self, export_id, batch_number):
        """Download specific batch file"""
        if export_id not in self.exports_storage:
            return None, None
        
        export_info = self.exports_storage[export_id]
        
        if "files_info" in export_info:
            for file_info in export_info["files_info"]:
                if file_info["batch_number"] == batch_number:
                    with open(file_info["file_path"], 'rb') as f:
                        return f.read(), file_info["file_name"]
        
        return None, None
    
    def _should_run_cleanup(self):
        """Check if cleanup should run based on time interval"""
        current_time = datetime.now()
        time_since_last_cleanup = (current_time - self.last_cleanup_time).total_seconds() / 60  # in minutes
        
        return time_since_last_cleanup >= self.cleanup_interval_minutes

    def _cleanup_old_exports(self):
        """Clean up old exports while respecting minimum age and time intervals"""
        if not self.exports_storage:
            return
            
        try:
            # Update last cleanup time
            self.last_cleanup_time = datetime.now()
            current_time = datetime.now()
            
            # Sort exports by creation time (newest first)
            exports_by_time = sorted(
                self.exports_storage.items(), 
                key=lambda x: x[1].get("created_at", datetime.min), 
                reverse=True
            )
            
            # Filter out exports that are too young to cleanup
            cleanable_exports = []
            protected_exports = []
            
            for export_id, export_info in exports_by_time:
                created_at = export_info.get("created_at", datetime.min)
                age_minutes = (current_time - created_at).total_seconds() / 60
                
                if age_minutes >= self.min_export_age_minutes:
                    cleanable_exports.append((export_id, export_info))
                else:
                    protected_exports.append((export_id, export_info))
            
            # Calculate how many we can keep
            total_exports = len(cleanable_exports) + len(protected_exports)
            
            # Only cleanup if we exceed the limit and have cleanable exports
            if total_exports > self.max_concurrent_exports and cleanable_exports:
                # Keep all protected exports + some cleanable ones
                keep_cleanable_count = max(0, self.max_concurrent_exports - len(protected_exports))
                exports_to_remove = cleanable_exports[keep_cleanable_count:]
                
                if exports_to_remove:
                    logger.info(f"Cleaning up {len(exports_to_remove)} old exports (total: {total_exports}, limit: {self.max_concurrent_exports}, protected: {len(protected_exports)})")
                    
                    for export_id, export_info in exports_to_remove:
                        age_minutes = (current_time - export_info.get('created_at', datetime.min)).total_seconds() / 60
                        logger.info(f"Cleaning up export {export_id} (age: {age_minutes:.1f} minutes)")
                        self._cleanup_export(export_id)
                        
        except Exception as e:
            logger.error(f"Error during old exports cleanup: {str(e)}")

    def _old_cleanup_old_exports(self):
        """Clean up old exports to maintain only the most recent ones"""
        try:
            # Get list of exports sorted by creation time (newest first)
            exports_by_time = sorted(
                self.exports_storage.items(), 
                key=lambda x: x[1].get("created_at", datetime.min), 
                reverse=True
            )
            
            # If we have more exports than the limit, clean up the old ones
            if len(exports_by_time) >= self.max_concurrent_exports:
                exports_to_remove = exports_by_time[self.max_concurrent_exports-1:]  # Keep one slot for new export
                
                for export_id, export_info in exports_to_remove:
                    logger.info(f"Cleaning up old export: {export_id}")
                    self._cleanup_export(export_id)
                    
        except Exception as e:
            logger.error(f"Error during old exports cleanup: {str(e)}")
    
    def _cleanup_export(self, export_id):
        """Clean up export files and data"""
        if export_id not in self.exports_storage:
            return
            
        export_info = self.exports_storage[export_id]
        files_cleaned = 0
        
        try:
            # Clean up temp files
            if "temp_files" in export_info:
                for temp_file in export_info["temp_files"]:
                    try:
                        if os.path.exists(temp_file):
                            os.remove(temp_file)
                            files_cleaned += 1
                            logger.debug(f"Deleted temp file: {temp_file}")
                    except Exception as e:
                        logger.warning(f"Failed to delete temp file {temp_file}: {str(e)}")
            
            # Clean up ZIP file
            if "zip_path" in export_info:
                try:
                    if os.path.exists(export_info["zip_path"]):
                        os.remove(export_info["zip_path"])
                        files_cleaned += 1
                        logger.debug(f"Deleted ZIP file: {export_info['zip_path']}")
                except Exception as e:
                    logger.warning(f"Failed to delete ZIP file {export_info['zip_path']}: {str(e)}")
                    
            # Clean up individual batch files (if stored separately)
            if "files_info" in export_info:
                for file_info in export_info["files_info"]:
                    file_path = file_info.get("file_path")
                    if file_path and os.path.exists(file_path):
                        try:
                            os.remove(file_path)
                            files_cleaned += 1
                            logger.debug(f"Deleted batch file: {file_path}")
                        except Exception as e:
                            logger.warning(f"Failed to delete batch file {file_path}: {str(e)}")
            
            # Remove from storage
            del self.exports_storage[export_id]
            
            logger.info(f"Cleaned up export {export_id}: {files_cleaned} files deleted")
            
        except Exception as e:
            logger.error(f"Error cleaning up export {export_id}: {str(e)}")
            # Still try to remove from storage even if file cleanup failed
            if export_id in self.exports_storage:
                del self.exports_storage[export_id]
    
    def cleanup_expired_exports(self):
        """Clean up all expired exports"""
        expired_exports = []
        current_time = datetime.now()
        
        for export_id, export_info in self.exports_storage.items():
            if current_time > export_info["expires_at"]:
                expired_exports.append(export_id)
        
        for export_id in expired_exports:
            logger.info(f"Cleaning up expired export: {export_id}")
            self._cleanup_export(export_id)
        
        logger.info(f"Cleaned up {len(expired_exports)} expired exports")
        return len(expired_exports)
    
    def force_cleanup_all_exports(self):
        """Force cleanup of all exports (admin function)"""
        export_ids = list(self.exports_storage.keys())
        
        for export_id in export_ids:
            self._cleanup_export(export_id)
        
        logger.info(f"Force cleaned up {len(export_ids)} exports")
        return len(export_ids)
    
    def get_storage_info(self):
        """Get information about current storage usage"""
        total_exports = len(self.exports_storage)
        total_files = 0
        total_size = 0
        
        for export_info in self.exports_storage.values():
            # Count temp files
            if "temp_files" in export_info:
                for temp_file in export_info["temp_files"]:
                    if os.path.exists(temp_file):
                        total_files += 1
                        try:
                            total_size += os.path.getsize(temp_file)
                        except:
                            pass
            
            # Count ZIP files
            if "zip_path" in export_info and os.path.exists(export_info["zip_path"]):
                total_files += 1
                try:
                    total_size += os.path.getsize(export_info["zip_path"])
                except:
                    pass
        
        return {
            "total_exports": total_exports,
            "total_files": total_files,
            "total_size_bytes": total_size,
            "total_size_mb": round(total_size / (1024 * 1024), 2),
            "max_concurrent_exports": self.max_concurrent_exports
        }
