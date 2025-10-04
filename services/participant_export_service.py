"""
Participant Export Service
Specialized service for exporting participants with complex joins and filters
"""
import uuid
from datetime import datetime, timedelta
import logging
import time
import pandas as pd
from typing import Dict, Any, Optional
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Database imports
from database.db_connection import get_db_connection
from database.participant_export_query_builder import participant_query_builder

# Configuration imports
from config.ybb_export_config import get_template, get_chunk_size, should_use_chunked_processing

logger = logging.getLogger(__name__)

class ParticipantExportService:
    """Specialized service for participant exports"""
    
    def __init__(self):
        self.exports_storage = {}  # In production, use Redis
        logger.info("✅ Participant Export Service initialized")
    
    def create_participant_export(self, program_id: int, filters: Dict[str, Any] = None, 
                                 force_chunking: bool = False, chunk_size: int = None) -> Dict[str, Any]:
        """
        Create participant export with complex filtering
        
        Args:
            export_config: Dictionary containing:
                - program_id (required): Program ID to export
                - template: Template name (basic, standard, detailed, complete)
                - filters: Dictionary of filters
                - filename: Optional custom filename
                - chunk_size: Optional chunk size override
                - force_chunking: Force multi-file output
        """
        start_time = time.time()
        export_id = f"exp_part_{uuid.uuid4().hex[:8]}"
        
        try:
            logger.info(f"PARTICIPANT_EXPORT_START | ID: {export_id}")
            
            # Validate required parameters
            if not program_id:
                return {
                    "status": "error",
                    "message": "program_id is required for participant export"
                }
            
            # Initialize filters if not provided
            if filters is None:
                filters = {}
            filters['program_id'] = program_id  # Ensure program_id is in filters
            
            # Create clean database filters (exclude data inclusion flags)
            db_filters = {k: v for k, v in filters.items() if k != 'with_essay'}
            logger.info(f"DB_FILTERS | ID: {export_id} | Database filters: {db_filters}")
            
            # Use provided parameters directly (no export_config needed)
            filename = f'participants_program_{program_id}_export'
            
            # Base participant fields (always included)
            template_fields = [
                'id', 'full_name', 'email', 'phone_number', 'gender', 'birthdate', 'nationality',
                'occupation', 'education_level', 'major', 'institution', 'organizations',
                'current_address', 'origin_address', 'instagram_account', 'emergency_contact_full',
                'contact_relation', 'disease_history', 'tshirt_size', 'category', 'experiences', 
                'achievements', 'knowledge_source', 'source_account_name', 'twibbon_link', 
                'requirement_link', 'ref_code_ambassador', 'score_total', 'form_status',
                'created_at', 'updated_at'
            ]
            
            # Conditionally add essay-related fields if with_essay=True
            if filters.get('with_essay', False):
                template_fields.extend(['subtheme_names', 'competition_categories'])
                logger.info(f"ESSAY_FIELDS_INCLUDED | ID: {export_id} | Essay data will be included")
            else:
                logger.info(f"ESSAY_FIELDS_EXCLUDED | ID: {export_id} | Essay data excluded for faster export")
            
            # Store template fields and filters for title section
            self.current_template_fields = template_fields
            self.current_filters = filters
            
            # Create human-readable headers mapping
            human_headers = self._get_human_readable_headers()
            
            logger.info(f"PARTICIPANT_TEMPLATE | ID: {export_id} | Using comprehensive template | Fields: {len(template_fields)}")
            
            # Get total record count using clean database filters
            total_records = self._get_participant_count(db_filters)
            logger.info(f"PARTICIPANT_COUNT | ID: {export_id} | Total records: {total_records}")
            
            if total_records == 0:
                return {
                    "status": "error",
                    "message": "No participants found matching the specified filters",
                    "filters_used": filters
                }
            
            # Determine processing strategy
            should_chunk = (
                force_chunking or 
                total_records > 15000  # Use hardcoded limit since we always use detailed export
            )
            
            # Set optimal chunk size
            if should_chunk and not chunk_size:
                chunk_size = min(4000, max(1000, total_records // 10))  # Optimal chunk size
            
            logger.info(
                f"PARTICIPANT_STRATEGY | ID: {export_id} | "
                f"Records: {total_records} | Chunked: {should_chunk} | "
                f"Chunk size: {chunk_size}"
            )
            
            # Create export based on strategy
            if should_chunk and total_records > (chunk_size or 5000):
                result = self._create_chunked_participant_export(
                    export_id, template_fields, filters, 
                    f"participants_program_{program_id}_export", chunk_size, total_records
                )
            else:
                result = self._create_single_participant_export(
                    export_id, template_fields, filters, 
                    f"participants_program_{program_id}_export", total_records
                )
            
            # Add timing and performance metrics
            total_time = time.time() - start_time
            if result["status"] == "success" and "performance_metrics" in result:
                result["performance_metrics"]["total_processing_time_seconds"] = round(total_time, 2)
                result["performance_metrics"]["database_query_efficiency"] = "participant_optimized"
                result["performance_metrics"]["memory_approach"] = "streaming" if should_chunk else "standard"
            
            logger.info(f"PARTICIPANT_EXPORT_COMPLETED | ID: {export_id} | Time: {total_time:.2f}s")
            
            return result
            
        except Exception as e:
            logger.error(f"PARTICIPANT_EXPORT_ERROR | ID: {export_id} | Error: {str(e)}", exc_info=True)
            return {
                "status": "error",
                "message": f"Participant export failed: {str(e)}",
                "export_id": export_id
            }
    
    def _get_participant_count(self, filters: Dict[str, Any]) -> int:
        """Get total participant count with filters"""
        try:
            query, params = participant_query_builder.build_participant_count_query(filters)
            
            with get_db_connection() as conn:
                import pandas as pd
                result = pd.read_sql(query, conn, params=params)
                return int(result.iloc[0]['total'])
                
        except Exception as e:
            logger.error(f"Participant count query failed: {str(e)}")
            return 0
    
    def _create_single_participant_export(self, export_id: str, template_fields: list, 
                                        filters: Dict, filename: str, total_records: int):
        """Create single file participant export"""
        try:
            start_time = time.time()
            
            # Build and execute query using clean database filters
            # Pass original filters for with_essay detection in joins
            query, params = participant_query_builder.build_participant_export_query(
                template_fields, 
                filters={**db_filters, 'with_essay': filters.get('with_essay', False)}
            )
            
            logger.info(f"PARTICIPANT_SINGLE_EXPORT | ID: {export_id} | Query built, executing...")
            
            with get_db_connection() as conn:
                import pandas as pd
                df = pd.read_sql(query, conn, params=params)
            
            processing_time = time.time() - start_time
            
            if df.empty:
                return {
                    "status": "error",
                    "message": "No participant data returned from query"
                }
            
            logger.info(f"PARTICIPANT_DATA_LOADED | ID: {export_id} | Records: {len(df)} | Time: {processing_time:.2f}s")
            
            # Create Excel file
            excel_start = time.time()
            file_content, file_info = self._create_excel_from_dataframe(
                df, template_fields, filename, export_id
            )
            excel_time = time.time() - excel_start
            
            # Store export
            export_data = {
                'export_id': export_id,
                'content': file_content,
                'filename': file_info['filename'],
                'created_at': datetime.now(),
                'expires_at': datetime.now() + timedelta(hours=24),
                'export_type': 'participants',
                'record_count': len(df),
                'filters': filters  # Store filters for title section
            }
            
            self.exports_storage[export_id] = export_data
            
            # Performance metrics
            total_time = processing_time + excel_time
            performance_metrics = {
                "total_processing_time_seconds": round(total_time, 2),
                "database_query_time_seconds": round(processing_time, 2),
                "excel_generation_time_seconds": round(excel_time, 2),
                "records_per_second": round(len(df) / total_time, 1) if total_time > 0 else 0,
                "memory_used_mb": self._estimate_memory_usage(df),
                "efficiency_metrics": {
                    "ms_per_record": round((total_time * 1000) / len(df), 2) if len(df) > 0 else 0,
                    "kb_per_record": round(len(file_content) / 1024 / len(df), 2) if len(df) > 0 else 0
                }
            }
            
            return {
                "status": "success",
                "export_strategy": "single_file",
                "data": {
                    "export_id": export_id,
                    "file_name": file_info['filename'],
                    "file_size": len(file_content),
                    "file_size_mb": round(len(file_content) / 1024 / 1024, 2),
                    "record_count": len(df),
                    "download_url": f"/api/ybb/export/{export_id}/download",
                    "expires_at": export_data['expires_at'].isoformat()
                },
                "performance_metrics": performance_metrics,
                "system_info": {
                    "export_type": "participants",
                    "template": "detailed",
                    "processing_mode": "participant_optimized",
                    "generated_at": datetime.now().isoformat()
                }
            }
            
        except Exception as e:
            logger.error(f"Single participant export failed: {str(e)}", exc_info=True)
            raise
    
    def _create_chunked_participant_export(self, export_id: str, template_fields: list, 
                                         filters: Dict, filename: str, chunk_size: int, total_records: int):
        """Create multi-file chunked participant export"""
        try:
            start_time = time.time()
            chunk_files = []
            chunk_number = 1
            
            logger.info(f"PARTICIPANT_CHUNKED_EXPORT | ID: {export_id} | Processing {total_records} records in chunks of {chunk_size}")
            
            # Process data in chunks using database pagination
            offset = 0
            total_processed = 0
            chunk_times = []
            
            while offset < total_records:
                chunk_start = time.time()
                
                logger.info(f"PARTICIPANT_CHUNK_PROCESSING | ID: {export_id} | Chunk {chunk_number} | Offset: {offset}")
                
                # Build query for this chunk
                # Use clean database filters with original with_essay flag for joins
                query, params = participant_query_builder.build_participant_export_query(
                    template_fields, 
                    filters=filters,
                    limit=chunk_size,
                    offset=offset
                )
                
                # Execute chunk query
                with get_db_connection() as conn:
                    import pandas as pd
                    chunk_df = pd.read_sql(query, conn, params=params)
                
                if chunk_df.empty:
                    logger.warning(f"Empty chunk at offset {offset}, stopping")
                    break
                
                # Create Excel for this chunk
                chunk_filename = f"{filename}_chunk_{chunk_number}_{export_id}"
                chunk_content, chunk_info = self._create_excel_from_dataframe(
                    chunk_df, template_fields, chunk_filename, export_id
                )
                
                chunk_time = time.time() - chunk_start
                chunk_times.append(chunk_time)
                
                # Store chunk info - use the chunk data returned from Excel creation
                chunk_data = {
                    "batch_number": chunk_number,
                    "file_name": chunk_info['filename'],
                    "file_content": chunk_content,
                    "file_size": len(chunk_content),
                    "record_count": len(chunk_df),
                    "record_range": f"{offset + 1}-{offset + len(chunk_df)}",
                    "processing_time_seconds": round(chunk_time, 2),
                    "records_per_second": round(len(chunk_df) / chunk_time, 1) if chunk_time > 0 else 0
                }
                
                chunk_files.append(chunk_data)
                total_processed += len(chunk_df)
                
                logger.info(
                    f"PARTICIPANT_CHUNK_COMPLETED | ID: {export_id} | Chunk {chunk_number} | "
                    f"Records: {len(chunk_df)} | Time: {chunk_time:.2f}s"
                )
                
                offset += chunk_size
                chunk_number += 1
            
            # Create ZIP archive
            zip_start = time.time()
            zip_content, zip_filename = self._create_zip_archive(chunk_files, filename, export_id)
            zip_time = time.time() - zip_start
            
            total_time = time.time() - start_time
            
            # Store export
            export_data = {
                'export_id': export_id,
                'content': zip_content,
                'filename': zip_filename,
                'chunk_files': chunk_files,
                'created_at': datetime.now(),
                'expires_at': datetime.now() + timedelta(hours=24),
                'export_type': 'participants',
                'record_count': total_processed,
                'filters': filters  # Store filters for title section
            }
            
            self.exports_storage[export_id] = export_data
            
            # Performance metrics
            performance_metrics = {
                "total_processing_time_seconds": round(total_time, 2),
                "data_preparation_time_seconds": 0.0,  # No data prep needed with direct queries
                "average_chunk_processing_time_seconds": round(sum(chunk_times) / len(chunk_times), 2) if chunk_times else 0,
                "compression_time_seconds": round(zip_time, 2),
                "total_records_per_second": round(total_processed / total_time, 1) if total_time > 0 else 0,
                "chunk_processing_times": [round(t, 2) for t in chunk_times],
                "efficiency_metrics": {
                    "processing_ms_per_record": round((total_time * 1000) / total_processed, 2) if total_processed > 0 else 0,
                    "kb_per_record_compressed": round(len(zip_content) / 1024 / total_processed, 2) if total_processed > 0 else 0,
                    "database_efficiency": "participant_streaming_chunks"
                }
            }
            
            # Calculate compression stats
            uncompressed_size = sum(chunk['file_size'] for chunk in chunk_files)
            compression_ratio = round((1 - len(zip_content) / uncompressed_size) * 100, 1) if uncompressed_size > 0 else 0
            
            return {
                "status": "success",
                "export_strategy": "multi_file",
                "data": {
                    "export_id": export_id,
                    "total_records": total_processed,
                    "total_files": len(chunk_files),
                    "individual_files": [
                        {
                            "batch_number": chunk["batch_number"],
                            "file_name": chunk["file_name"],
                            "file_size": chunk["file_size"],
                            "record_count": chunk["record_count"],
                            "record_range": chunk["record_range"],
                            "processing_time_seconds": chunk["processing_time_seconds"],
                            "records_per_second": chunk["records_per_second"]
                        } for chunk in chunk_files
                    ],
                    "archive_info": {
                        "filename": zip_filename,
                        "compressed_size": len(zip_content),
                        "uncompressed_size": uncompressed_size,
                        "compression_ratio": f"{compression_ratio}%",
                        "compression_time_seconds": round(zip_time, 2)
                    }
                },
                "performance_metrics": performance_metrics,
                "download_url": f"/api/ybb/export/{export_id}/download"
            }
            
        except Exception as e:
            logger.error(f"Chunked participant export failed: {str(e)}", exc_info=True)
            raise
    
    def _create_excel_from_dataframe(self, df, template_fields, filename, export_id):
        """Create Excel file from pandas DataFrame with proper formatting"""
        try:
            logger.info(f"EXCEL_CREATE_START | ID: {export_id} | Rows: {len(df)} | Filename: {filename}")
            
            # Comprehensive Excel-safe data cleaning
            df = self._sanitize_dataframe_for_excel(df, export_id)
            
            logger.info(f"EXCEL_DATA_SANITIZED | ID: {export_id} | DataFrame fully sanitized for Excel")
            
            # Create workbook with title section
            wb = Workbook()
            ws = wb.active
            ws.title = "Participants"
            
            # Add title section
            current_row = self._add_title_section(ws, template_fields, export_id)
            
            # Apply human-readable headers
            human_headers = self._get_human_readable_headers()
            headers = [human_headers.get(field, field) for field in template_fields]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Add data rows with proper formatting (start after title section and headers)
            data_start_row = current_row + 1
            for row_idx, (_, row) in enumerate(df.iterrows(), data_start_row):
                for col_idx, field in enumerate(template_fields, 1):
                    value = row.get(field, '')
                    
                    # Handle different data types (data is already sanitized at DataFrame level)
                    if value is None or pd.isna(value):
                        value = ''
                    elif isinstance(value, datetime):
                        try:
                            value = value.strftime('%Y-%m-%d %H:%M:%S') if value else ''
                        except:
                            value = str(value) if value else ''
                    elif str(value).lower() in ['nan', 'nat']:
                        value = ''
                    elif isinstance(value, (int, float)) and field in ['form_status', 'payment_status', 'general_status', 'document_status']:
                        # Convert status numbers to readable text
                        if field == 'form_status':
                            status_map = {0: 'Not Started', 1: 'In Progress', 2: 'Submitted'}
                            try:
                                value = status_map.get(int(value), str(value))
                            except (ValueError, TypeError):
                                value = str(value)
                        else:
                            value = str(value)
                    # Value is already sanitized from DataFrame cleaning
                    
                    ws.cell(row=row_idx, column=col_idx, value=str(value))
            
            # Auto-adjust column widths (safer approach)
            from openpyxl.utils import get_column_letter
            from openpyxl.cell.cell import MergedCell
            
            for col_num in range(1, ws.max_column + 1):
                max_length = 0
                column_letter = get_column_letter(col_num)
                
                # Check all cells in this column, skipping merged cells
                for row_num in range(1, ws.max_row + 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    if not isinstance(cell, MergedCell) and cell.value is not None:
                        try:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                        except:
                            pass
                
                # Set reasonable column width with special handling for ID column
                if col_num == 1:  # ID column should be smaller
                    adjusted_width = min(max(max_length + 1, 6), 12)  # ID: min 6, max 12
                else:
                    adjusted_width = min(max(max_length + 2, 12), 50)  # Others: min 12, max 50
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save to BytesIO
            file_buffer = BytesIO()
            wb.save(file_buffer)
            file_content = file_buffer.getvalue()
            file_buffer.close()
            
            # Generate descriptive filename with timestamp
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            descriptive_filename = self._generate_descriptive_filename(export_id, len(df), timestamp)
            final_filename = f"{descriptive_filename}.xlsx"
            
            return file_content, {
                'filename': final_filename,
                'size': len(file_content),
                'records': len(df)
            }
            
        except Exception as e:
            logger.error(f"Excel creation failed: {str(e)}", exc_info=True)
            raise
    
    def _sanitize_dataframe_for_excel(self, df, export_id):
        """Comprehensive DataFrame sanitization for Excel compatibility"""
        import re
        
        logger.info(f"SANITIZE_START | ID: {export_id} | Original shape: {df.shape}")
        
        # Make a copy to avoid modifying original
        df_clean = df.copy()
        
        # Define problematic characters that cause Excel issues
        problematic_chars = [
            '\x00', '\x01', '\x02', '\x03', '\x04', '\x05', '\x06', '\x07', '\x08',
            '\x0b', '\x0c', '\x0e', '\x0f', '\x10', '\x11', '\x12', '\x13', '\x14',
            '\x15', '\x16', '\x17', '\x18', '\x19', '\x1a', '\x1b', '\x1c', '\x1d',
            '\x1e', '\x1f'
        ]
        
        cleaned_cells = 0
        
        for col in df_clean.columns:
            if df_clean[col].dtype == 'object':  # String/text columns
                original_values = df_clean[col].copy()
                
                # Convert to string and handle nulls
                df_clean[col] = df_clean[col].astype(str)
                df_clean[col] = df_clean[col].replace(['nan', 'None', 'NaT'], '', regex=False)
                
                # Remove problematic control characters
                for char in problematic_chars:
                    df_clean[col] = df_clean[col].str.replace(char, '', regex=False)
                
                # Replace line breaks with spaces
                df_clean[col] = df_clean[col].str.replace('\n', ' ', regex=False)
                df_clean[col] = df_clean[col].str.replace('\r', ' ', regex=False)
                df_clean[col] = df_clean[col].str.replace('\t', ' ', regex=False)
                
                # Remove excessive whitespace
                df_clean[col] = df_clean[col].str.replace(r'\s+', ' ', regex=True)
                df_clean[col] = df_clean[col].str.strip()
                
                # Enforce Excel cell character limit (32,767 characters)
                df_clean[col] = df_clean[col].str.slice(0, 32767)
                
                # Remove non-printable characters except common ones
                df_clean[col] = df_clean[col].str.replace(r'[^\x20-\x7E\x80-\xFF]', '', regex=True)
                
                # Count cleaned cells
                changes = (original_values.astype(str) != df_clean[col]).sum()
                cleaned_cells += changes
                
                if changes > 0:
                    logger.info(f"SANITIZE_COLUMN | ID: {export_id} | Column: {col} | Cleaned: {changes} cells")
        
        # Handle numeric columns - convert inf and -inf to None
        try:
            numeric_cols = df_clean.select_dtypes(include=['int64', 'float64', 'int32', 'float32']).columns
            for col in numeric_cols:
                df_clean[col] = df_clean[col].replace([float('inf'), float('-inf')], None)
        except Exception as e:
            logger.warning(f"SANITIZE_NUMERIC_SKIP | ID: {export_id} | Error: {e}")
        
        logger.info(f"SANITIZE_COMPLETE | ID: {export_id} | Total cells cleaned: {cleaned_cells}")
        
        return df_clean
    
    def _get_human_readable_headers(self):
        """Return mapping of field names to human-readable headers"""
        return {
            'id': 'ID',
            'full_name': 'Full Name',
            'email': 'Email Address',
            'phone_number': 'Phone Number',
            'gender': 'Gender',
            'birthdate': 'Date of Birth',
            'nationality': 'Nationality',
            'occupation': 'Occupation',
            'education_level': 'Education Level',
            'major': 'Major/Field of Study',
            'institution': 'Institution/University',
            'organizations': 'Organizations',
            'current_address': 'Current Address',
            'origin_address': 'Origin Address',
            'instagram_account': 'Instagram Account',
            'emergency_contact_full': 'Emergency Contact',
            'contact_relation': 'Emergency Contact Relation',
            'disease_history': 'Medical History',
            'tshirt_size': 'T-Shirt Size',
            'category': 'Participation Category',
            'experiences': 'Relevant Experiences',
            'achievements': 'Achievements',
            'knowledge_source': 'How Did You Know About Us',
            'source_account_name': 'Source Account Name',
            'twibbon_link': 'Twibbon Link',
            'requirement_link': 'Requirement Submission Link',
            'ref_code_ambassador': 'Ambassador Reference Code',
            'score_total': 'Total Score',
            'form_status': 'Application Status',
            'subtheme_names': 'Selected Sub-themes',
            'competition_categories': 'Competition Categories',
            'created_at': 'Registration Date',
            'updated_at': 'Last Updated'
        }
    
    def _add_title_section(self, worksheet, template_fields, export_id):
        """Add title section to Excel worksheet"""
        from datetime import datetime
        
        # Main title
        title_cell = worksheet.cell(row=1, column=1, value="Japan Youth Summit 2025 Participant Data")
        title_cell.font = Font(size=16, bold=True, color="2F5597")
        worksheet.merge_cells('A1:E1')
        
        # Generation date
        generation_date = datetime.now().strftime('%A, %B %d, %Y %H:%M')
        date_cell = worksheet.cell(row=3, column=1, value=f"Generated on {generation_date}")
        date_cell.font = Font(size=11, italic=True, color="666666")
        worksheet.merge_cells('A3:D3')
        
        # Get filter information from stored export data
        filter_info = self._get_filter_description(export_id)
        if filter_info:
            filter_cell = worksheet.cell(row=5, column=1, value=f"Filters: {filter_info}")
            filter_cell.font = Font(size=10, color="666666")
            worksheet.merge_cells('A5:F5')
            header_row = 7
        else:
            header_row = 6
        
        # Add spacing
        return header_row
    
    def _get_filter_description(self, export_id):
        """Generate human-readable filter description"""
        # Use current filters (more reliable than stored data)
        filters = getattr(self, 'current_filters', {})
        
        # Fallback to stored export data if current filters not available
        if not filters:
            export_data = self.exports_storage.get(export_id, {})
            filters = export_data.get('filters', {})
        
        # Remove with_essay=False from filters for description (it's not a real filter)
        display_filters = {k: v for k, v in filters.items() if not (k == 'with_essay' and v is False)}
        
        descriptions = []
        
        # Category filter
        if display_filters.get('category'):
            if isinstance(display_filters['category'], list):
                descriptions.append(f"Categories: {', '.join(display_filters['category'])}")
            else:
                descriptions.append(f"Category: {display_filters['category']}")
        
        # Form status filter
        if filters.get('form_status') is not None:
            status_map = {0: 'Not Started', 1: 'In Progress', 2: 'Submitted'}
            if isinstance(filters['form_status'], list):
                statuses = [status_map.get(s, str(s)) for s in filters['form_status']]
                descriptions.append(f"Application Status: {', '.join(statuses)}")
            else:
                status = status_map.get(filters['form_status'], str(filters['form_status']))
                descriptions.append(f"Application Status: {status}")
        
        # Payment filters
        if filters.get('payment_done'):
            descriptions.append("With successful payments")
        elif filters.get('payment_status') is not None:
            payment_map = {0: 'Pending', 1: 'Processing', 2: 'Completed', 3: 'Failed'}
            if isinstance(filters['payment_status'], list):
                statuses = [payment_map.get(s, str(s)) for s in filters['payment_status']]
                descriptions.append(f"Payment Status: {', '.join(statuses)}")
            else:
                status = payment_map.get(filters['payment_status'], str(filters['payment_status']))
                descriptions.append(f"Payment Status: {status}")
        
        # Essay data inclusion
        if filters.get('with_essay'):
            descriptions.append("Including essay data and sub-themes")
        else:
            descriptions.append("Basic participant data only")
        
        # Essay count filters (if any)
        if filters.get('essay_count_min'):
            descriptions.append(f"Minimum {filters['essay_count_min']} essays")
        
        # Score filters
        if filters.get('score_min') or filters.get('score_max'):
            score_parts = []
            if filters.get('score_min'):
                score_parts.append(f"score ≥ {filters['score_min']}")
            if filters.get('score_max'):
                score_parts.append(f"score ≤ {filters['score_max']}")
            descriptions.append(f"Score range: {', '.join(score_parts)}")
        
        # Date filters
        if filters.get('created_at_min') or filters.get('created_at_max'):
            date_parts = []
            if filters.get('created_at_min'):
                date_parts.append(f"from {filters['created_at_min']}")
            if filters.get('created_at_max'):
                date_parts.append(f"to {filters['created_at_max']}")
            descriptions.append(f"Registration dates: {' '.join(date_parts)}")
        
        # Education level filter
        if filters.get('education_level'):
            if isinstance(filters['education_level'], list):
                descriptions.append(f"Education: {', '.join(filters['education_level'])}")
            else:
                descriptions.append(f"Education: {filters['education_level']}")
        
        # Nationality filter
        if filters.get('nationality'):
            if isinstance(filters['nationality'], list):
                descriptions.append(f"Nationality: {', '.join(filters['nationality'])}")
            else:
                descriptions.append(f"Nationality: {filters['nationality']}")
        
        # Don't add 'All participants' automatically - be more specific
        if len(descriptions) == 1 and 'Basic participant data only' in descriptions[0]:
            return "All participants - Basic data export"
        elif len(descriptions) == 1 and 'Including essay data' in descriptions[0]:
            return "All participants - Including essay data and sub-themes"
        elif not descriptions:
            return "All participants - Basic data export"
        
        return "; ".join(descriptions)
    
    def _generate_descriptive_filename(self, export_id, record_count, timestamp):
        """Generate descriptive filename based on export content"""
        # Get export data for context
        export_data = self.exports_storage.get(export_id, {})
        filters = export_data.get('filters', {})
        
        # Base name
        parts = ["YBS2025_Participants"]
        
        # Add program info if available
        program_id = filters.get('program_id')
        if program_id:
            parts.append(f"Program{program_id}")
        
        # Add filter descriptions
        if filters.get('category'):
            if isinstance(filters['category'], list):
                parts.append(f"Cat{'_'.join(filters['category'])}")
            else:
                parts.append(f"Cat{filters['category']}")
        
        if filters.get('form_status') is not None:
            status_map = {0: 'NotStarted', 1: 'InProgress', 2: 'Submitted'}
            if isinstance(filters['form_status'], list):
                statuses = [status_map.get(s, f'Status{s}') for s in filters['form_status']]
                parts.append('_'.join(statuses))
            else:
                status = status_map.get(filters['form_status'], f'Status{filters["form_status"]}')
                parts.append(status)
        
        if filters.get('payment_done'):
            parts.append("Paid")
        
        if filters.get('with_essay'):
            parts.append("WithEssays")
        
        # Add record count
        parts.append(f"{record_count}recs")
        
        # Add timestamp
        parts.append(timestamp)
        
        # Join with underscores and clean up
        filename = "_".join(parts)
        # Remove invalid filename characters
        invalid_chars = '<>:"/\\|?*'
        for char in invalid_chars:
            filename = filename.replace(char, '')
        
        return filename
    
    def _create_zip_archive(self, chunk_files, base_filename, export_id):
        """Create ZIP archive from multiple chunk files"""
        try:
            import zipfile
            
            zip_buffer = BytesIO()
            timestamp = datetime.now().strftime('%d-%m-%Y_%H%M%S')
            zip_filename = f"{base_filename}_{export_id}_{timestamp}.zip"
            
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED, compresslevel=6) as zipf:
                for chunk in chunk_files:
                    zipf.writestr(chunk['file_name'], chunk['file_content'])
            
            zip_content = zip_buffer.getvalue()
            zip_buffer.close()
            
            return zip_content, zip_filename
            
        except Exception as e:
            logger.error(f"ZIP creation failed: {str(e)}", exc_info=True)
            raise
    
    def _estimate_memory_usage(self, df):
        """Estimate memory usage of DataFrame"""
        try:
            return round(df.memory_usage(deep=True).sum() / 1024 / 1024, 2)
        except:
            return 0.0
    
    def download_export(self, export_id):
        """Download export file"""
        try:
            if export_id not in self.exports_storage:
                return None, None
            
            export_data = self.exports_storage[export_id]
            
            # Check expiration
            if datetime.now() > export_data['expires_at']:
                del self.exports_storage[export_id]
                return None, None
            
            return export_data['content'], export_data['filename']
            
        except Exception as e:
            logger.error(f"Download failed for export {export_id}: {str(e)}")
            return None, None

# Global participant export service instance
participant_export_service = ParticipantExportService()