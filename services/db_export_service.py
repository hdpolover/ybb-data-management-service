"""
Database Export Service - High-performance direct database export
Eliminates JSON payload limitations by processing data directly from database
"""
import uuid
from datetime import datetime, timedelta
import os
import tempfile
import logging
import time
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Database imports
from database.db_connection import get_db_connection, execute_chunked_query, get_table_count
from database.query_builder import build_export_query, build_count_query

# Configuration imports
from config.ybb_export_config import get_template, get_chunk_size, should_use_chunked_processing

# Existing file management
from utils.file_manager import ExportFileManager

logger = logging.getLogger(__name__)

class DatabaseExportService:
    """High-performance export service using direct database access"""
    
    def __init__(self):
        self.file_manager = ExportFileManager()
        self.exports_storage = {}  # In production, use Redis
        self.temp_dir = tempfile.gettempdir()
        
        logger.info("✅ Database Export Service initialized (lazy DB connection)")
    
    def create_database_export(self, export_config):
        """
        Create export by querying database directly
        Much faster than JSON payload approach
        """
        start_time = time.time()
        export_id = f"exp_db_{uuid.uuid4().hex[:8]}"
        
        try:
            logger.info(f"DB_EXPORT_START | ID: {export_id} | Type: {export_config.get('export_type')}")
            
            # Extract configuration
            export_type = export_config.get('export_type', 'participants')
            template = export_config.get('template', 'standard')
            filters = export_config.get('filters', {})
            filename = export_config.get('filename', f'{export_type}_export')
            chunk_size = export_config.get('chunk_size')
            force_chunking = export_config.get('force_chunking', False)
            total_records = export_config.get('total_records', 0)
            
            # Get template configuration
            template_config = get_template(export_type, template)
            if not template_config:
                return {
                    "status": "error",
                    "message": f"Invalid template '{template}' for export type '{export_type}'"
                }
            
            logger.info(f"DB_EXPORT_TEMPLATE | ID: {export_id} | Template: {template} | Fields: {len(template_config.get('fields', []))}")
            
            # Determine processing strategy
            should_chunk = (
                force_chunking or 
                should_use_chunked_processing(total_records, template_config) or
                total_records > template_config.get('max_records_single_file', 15000)
            )
            
            # Set optimal chunk size
            if should_chunk and not chunk_size:
                chunk_size = get_chunk_size(export_type, template, total_records)
            
            logger.info(
                f"DB_EXPORT_STRATEGY | ID: {export_id} | "
                f"Records: {total_records} | Chunked: {should_chunk} | "
                f"Chunk size: {chunk_size}"
            )
            
            # Create export based on strategy
            if should_chunk and total_records > chunk_size:
                result = self._create_chunked_database_export(
                    export_id, export_type, template_config, filters, 
                    filename, chunk_size, total_records
                )
            else:
                result = self._create_single_database_export(
                    export_id, export_type, template_config, filters, 
                    filename, total_records
                )
            
            # Add timing and performance metrics
            total_time = time.time() - start_time
            if result["status"] == "success" and "performance_metrics" in result:
                result["performance_metrics"]["total_processing_time_seconds"] = round(total_time, 2)
                result["performance_metrics"]["database_query_efficiency"] = "direct"
                result["performance_metrics"]["memory_approach"] = "streaming" if should_chunk else "standard"
            
            logger.info(f"DB_EXPORT_COMPLETED | ID: {export_id} | Time: {total_time:.2f}s")
            
            return result
            
        except Exception as e:
            logger.error(f"DB_EXPORT_ERROR | ID: {export_id} | Error: {str(e)}", exc_info=True)
            return {
                "status": "error",
                "message": f"Database export failed: {str(e)}",
                "export_id": export_id
            }
    
    def _create_single_database_export(self, export_id, export_type, template_config, 
                                     filters, filename, total_records):
        """Create single file export from database"""
        try:
            start_time = time.time()
            
            # Build SQL query
            from database.query_builder import query_builder
            query, params = query_builder.build_export_query(
                export_type, 
                template_config['fields'], 
                filters=filters
            )
            
            logger.info(f"DB_SINGLE_EXPORT | ID: {export_id} | Query built, executing...")
            
            # Execute query and get data
            with get_db_connection() as conn:
                import pandas as pd
                df = pd.read_sql(query, conn, params=params)
            
            processing_time = time.time() - start_time
            
            if df.empty:
                return {
                    "status": "error",
                    "message": "No data returned from database query"
                }
            
            logger.info(f"DB_DATA_LOADED | ID: {export_id} | Records: {len(df)} | Time: {processing_time:.2f}s")
            
            # Create Excel file
            excel_start = time.time()
            file_content, file_info = self._create_excel_from_dataframe(
                df, template_config, filename, export_id
            )
            excel_time = time.time() - excel_start
            
            # Store export
            export_data = {
                'export_id': export_id,
                'content': file_content,
                'filename': file_info['filename'],
                'created_at': datetime.now(),
                'expires_at': datetime.now() + timedelta(hours=24),
                'export_type': export_type,
                'record_count': len(df)
            }
            
            self.exports_storage[export_id] = export_data
            
            # Performance metrics
            total_time = processing_time + excel_time
            performance_metrics = {
                "total_processing_time_seconds": round(total_time, 2),
                "database_query_time_seconds": round(processing_time, 2),
                "excel_generation_time_seconds": round(excel_time, 2),
                "records_per_second": round(len(df) / total_time, 1) if total_time > 0 else 0,
                "memory_used_mb": self._estimate_memory_usage(df),
                "efficiency_metrics": {
                    "ms_per_record": round((total_time * 1000) / len(df), 2) if len(df) > 0 else 0,
                    "kb_per_record": round(len(file_content) / 1024 / len(df), 2) if len(df) > 0 else 0
                }
            }
            
            return {
                "status": "success",
                "export_strategy": "single_file",
                "data": {
                    "export_id": export_id,
                    "file_name": file_info['filename'],
                    "file_size": len(file_content),
                    "file_size_mb": round(len(file_content) / 1024 / 1024, 2),
                    "record_count": len(df),
                    "download_url": f"/api/ybb/export/{export_id}/download",
                    "expires_at": export_data['expires_at'].isoformat()
                },
                "performance_metrics": performance_metrics,
                "system_info": {
                    "export_type": export_type,
                    "template": template_config.get('name', 'unknown'),
                    "processing_mode": "database_direct",
                    "generated_at": datetime.now().isoformat()
                }
            }
            
        except Exception as e:
            logger.error(f"Single database export failed: {str(e)}", exc_info=True)
            raise
    
    def _create_chunked_database_export(self, export_id, export_type, template_config, 
                                      filters, filename, chunk_size, total_records):
        """Create multi-file chunked export from database"""
        try:
            start_time = time.time()
            chunk_files = []
            chunk_number = 1
            
            logger.info(f"DB_CHUNKED_EXPORT | ID: {export_id} | Processing {total_records} records in chunks of {chunk_size}")
            
            # Build base query for chunked processing
            base_query, params = build_export_query(
                export_type, 
                template_config['fields'], 
                filters=filters,
                order_by=template_config.get('order_by', 'id ASC')
            )
            
            # Process data in chunks using database pagination
            offset = 0
            total_processed = 0
            chunk_times = []
            
            while offset < total_records:
                chunk_start = time.time()
                
                # Add LIMIT/OFFSET to query
                chunk_query = f"{base_query} LIMIT {chunk_size} OFFSET {offset}"
                
                logger.info(f"DB_CHUNK_PROCESSING | ID: {export_id} | Chunk {chunk_number} | Offset: {offset}")
                
                # Execute chunk query
                with get_db_connection() as conn:
                    import pandas as pd
                    chunk_df = pd.read_sql(chunk_query, conn, params=params)
                
                if chunk_df.empty:
                    logger.warning(f"Empty chunk at offset {offset}, stopping")
                    break
                
                # Create Excel for this chunk
                chunk_filename = f"{filename}_chunk_{chunk_number}_{export_id}"
                file_content, file_info = self._create_excel_from_dataframe(
                    chunk_df, template_config, chunk_filename, export_id
                )
                
                chunk_time = time.time() - chunk_start
                chunk_times.append(chunk_time)
                
                # Store chunk info
                chunk_info = {
                    "batch_number": chunk_number,
                    "file_name": file_info['filename'],
                    "file_content": file_content,
                    "file_size": len(file_content),
                    "record_count": len(chunk_df),
                    "record_range": f"{offset + 1}-{offset + len(chunk_df)}",
                    "processing_time_seconds": round(chunk_time, 2),
                    "records_per_second": round(len(chunk_df) / chunk_time, 1) if chunk_time > 0 else 0
                }
                
                chunk_files.append(chunk_info)
                total_processed += len(chunk_df)
                
                logger.info(
                    f"DB_CHUNK_COMPLETED | ID: {export_id} | Chunk {chunk_number} | "
                    f"Records: {len(chunk_df)} | Time: {chunk_time:.2f}s"
                )
                
                offset += chunk_size
                chunk_number += 1
            
            # Create ZIP archive
            zip_start = time.time()
            zip_content, zip_filename = self._create_zip_archive(chunk_files, filename, export_id)
            zip_time = time.time() - zip_start
            
            total_time = time.time() - start_time
            
            # Store export
            export_data = {
                'export_id': export_id,
                'content': zip_content,
                'filename': zip_filename,
                'chunk_files': chunk_files,
                'created_at': datetime.now(),
                'expires_at': datetime.now() + timedelta(hours=24),
                'export_type': export_type,
                'record_count': total_processed
            }
            
            self.exports_storage[export_id] = export_data
            
            # Performance metrics
            performance_metrics = {
                "total_processing_time_seconds": round(total_time, 2),
                "data_preparation_time_seconds": 0.0,  # No data prep needed with DB direct
                "average_chunk_processing_time_seconds": round(sum(chunk_times) / len(chunk_times), 2),
                "compression_time_seconds": round(zip_time, 2),
                "total_records_per_second": round(total_processed / total_time, 1) if total_time > 0 else 0,
                "chunk_processing_times": [round(t, 2) for t in chunk_times],
                "efficiency_metrics": {
                    "processing_ms_per_record": round((total_time * 1000) / total_processed, 2),
                    "kb_per_record_compressed": round(len(zip_content) / 1024 / total_processed, 2),
                    "database_efficiency": "streaming_chunks"
                }
            }
            
            # Calculate compression stats
            uncompressed_size = sum(chunk['file_size'] for chunk in chunk_files)
            compression_ratio = round((1 - len(zip_content) / uncompressed_size) * 100, 1) if uncompressed_size > 0 else 0
            
            return {
                "status": "success",
                "export_strategy": "multi_file",
                "data": {
                    "export_id": export_id,
                    "total_records": total_processed,
                    "total_files": len(chunk_files),
                    "individual_files": [
                        {
                            "batch_number": chunk["batch_number"],
                            "file_name": chunk["file_name"],
                            "file_size": chunk["file_size"],
                            "record_count": chunk["record_count"],
                            "record_range": chunk["record_range"],
                            "processing_time_seconds": chunk["processing_time_seconds"],
                            "records_per_second": chunk["records_per_second"]
                        } for chunk in chunk_files
                    ],
                    "archive_info": {
                        "filename": zip_filename,
                        "compressed_size": len(zip_content),
                        "uncompressed_size": uncompressed_size,
                        "compression_ratio": f"{compression_ratio}%",
                        "compression_time_seconds": round(zip_time, 2)
                    }
                },
                "performance_metrics": performance_metrics,
                "download_url": f"/api/ybb/export/{export_id}/download"
            }
            
        except Exception as e:
            logger.error(f"Chunked database export failed: {str(e)}", exc_info=True)
            raise
    
    def _create_excel_from_dataframe(self, df, template_config, filename, export_id):
        """Create Excel file from pandas DataFrame"""
        try:
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Export Data"
            
            # Apply headers
            headers = template_config.get('headers', template_config['fields'])
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Add data rows
            for row_idx, (_, row) in enumerate(df.iterrows(), 2):
                for col_idx, field in enumerate(template_config['fields'], 1):
                    value = row.get(field, '')
                    # Handle different data types
                    if value is None or (hasattr(value, '__iter__') and str(value).lower() == 'nan'):
                        value = ''
                    elif isinstance(value, datetime):
                        value = value.strftime('%Y-%m-%d %H:%M:%S') if value else ''
                    
                    ws.cell(row=row_idx, column=col_idx, value=str(value))
            
            # Auto-adjust column widths (basic implementation)
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save to BytesIO
            file_buffer = BytesIO()
            wb.save(file_buffer)
            file_content = file_buffer.getvalue()
            file_buffer.close()
            
            # Generate filename with timestamp
            timestamp = datetime.now().strftime('%d-%m-%Y_%H%M%S')
            final_filename = f"{filename}_{timestamp}.xlsx"
            
            return file_content, {
                'filename': final_filename,
                'size': len(file_content),
                'records': len(df)
            }
            
        except Exception as e:
            logger.error(f"Excel creation failed: {str(e)}", exc_info=True)
            raise
    
    def _create_zip_archive(self, chunk_files, base_filename, export_id):
        """Create ZIP archive from multiple chunk files"""
        try:
            import zipfile
            
            zip_buffer = BytesIO()
            timestamp = datetime.now().strftime('%d-%m-%Y_%H%M%S')
            zip_filename = f"{base_filename}_{export_id}_{timestamp}.zip"
            
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED, compresslevel=6) as zipf:
                for chunk in chunk_files:
                    zipf.writestr(chunk['file_name'], chunk['file_content'])
            
            zip_content = zip_buffer.getvalue()
            zip_buffer.close()
            
            return zip_content, zip_filename
            
        except Exception as e:
            logger.error(f"ZIP creation failed: {str(e)}", exc_info=True)
            raise
    
    def _estimate_memory_usage(self, df):
        """Estimate memory usage of DataFrame"""
        try:
            return round(df.memory_usage(deep=True).sum() / 1024 / 1024, 2)
        except:
            return 0.0
    
    def get_export_record_count(self, export_type, filters=None):
        """Get total record count for export without loading data"""
        try:
            from database.query_builder import query_builder
            query, params = query_builder.build_count_query(export_type, filters)
            
            with get_db_connection() as conn:
                import pandas as pd
                result = pd.read_sql(query, conn, params=params)
                return int(result.iloc[0]['total'])
                
        except Exception as e:
            logger.error(f"Record count query failed: {str(e)}")
            return 0
    
    def get_export_preview(self, export_type, template, filters=None, limit=100):
        """Get preview of export data"""
        try:
            template_config = get_template(export_type, template)
            if not template_config:
                return []
            
            from database.query_builder import query_builder
            query, params = query_builder.build_export_query(
                export_type, 
                template_config['fields'], 
                filters=filters, 
                limit=limit
            )
            
            with get_db_connection() as conn:
                import pandas as pd
                df = pd.read_sql(query, conn, params=params)
                
                # Convert to list of dictionaries
                return df.to_dict('records')
                
        except Exception as e:
            logger.error(f"Preview query failed: {str(e)}")
            return []
    
    def estimate_export_metrics(self, export_type, total_records):
        """Estimate processing time and file size"""
        try:
            # Basic estimates based on record count
            # These can be refined based on actual performance data
            
            estimated_time_seconds = max(5, total_records * 0.0008)  # ~0.8ms per record
            estimated_size_mb = total_records * 0.0002  # ~0.2KB per record
            
            return {
                "estimated_processing_time_seconds": round(estimated_time_seconds, 1),
                "estimated_file_size_mb": round(estimated_size_mb, 2),
                "recommended_chunk_size": 4000 if total_records > 10000 else None,
                "recommended_processing": "chunked" if total_records > 5000 else "single_file"
            }
            
        except Exception as e:
            logger.error(f"Estimation failed: {str(e)}")
            return {}
    
    def get_available_tables_info(self):
        """Get information about available tables and fields"""
        try:
            from database.query_builder import query_builder
            
            table_info = {}
            for table_name, config in query_builder.TABLE_CONFIGS.items():
                table_info[table_name] = {
                    "table": config['table'],
                    "available_fields": config['allowed_fields'],
                    "primary_key": config['primary_key'],
                    "supports_joins": bool(config.get('joins', {}))
                }
            
            return table_info
            
        except Exception as e:
            logger.error(f"Table info retrieval failed: {str(e)}")
            return {}
    
    def download_export(self, export_id, file_type="single"):
        """Download export file (compatible with existing interface)"""
        try:
            if export_id not in self.exports_storage:
                return None, None
            
            export_data = self.exports_storage[export_id]
            
            # Check expiration
            if datetime.now() > export_data['expires_at']:
                del self.exports_storage[export_id]
                return None, None
            
            return export_data['content'], export_data['filename']
            
        except Exception as e:
            logger.error(f"Download failed for export {export_id}: {str(e)}")
            return None, None